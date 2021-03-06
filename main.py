
import sqlite3
#
#

import PyQt5
from PyQt5 import QtWidgets, QtCore, QtGui
from PyQt5.QtGui import QCursor
from PyQt5.QtWidgets import QVBoxLayout, QTableWidgetItem, QLabel, QMessageBox
import sys
from UI import Ui_MainWindow
from UI_Choose import Ui_Form
from UI_dialog import Ui_Dialog
from UI_bu_help import Ui_Dialog_help
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import (NavigationToolbar2QT as NavigationToolbar)
import matplotlib.pyplot as plt
import math
import numpy as np
from scipy import interpolate
import openpyxl
from openpyxl.drawing.image import Image

from prettytable import PrettyTable
#import matplotlib.patches
# GLOBAL
# Wing
# from os.path import expanduser

l = 0
S = 0
b = 0
b0 = 0
bk = 0
n = 0
c_ = 0
xc_ = 0
f_ = 0
a0 = 0
xf_ = 0
xc_degree = 0
x_degree = 0
lamda = 0
Sf = 0
Sf_ = 0
Sgd = 0
Sgd_ = 0
Sgsh = 0
Sgh_ = 0
S_ = 0
lamdaef = 0
caya = 0
Sobd = 0
Sobd_ = 0
xtau_ = 0
cmo = 0
h = 0

# Flap and Mulard
lzak = 0
bzak_ = 0
Sobzak = 0
Sobzak_ = 0
deltavzl = 0
deltapos = 0
xshzak = 0
bsrzak = 0
hvzl = 0
hpos = 0
Sobpr = 0
Sobpr_ = 0
Flap_type = 'Исходное крыло'
dCxomax = 0

# Tail
bgo = 0
cgo_ = 0
lgo = 0
Sgo = 0
lamdago = 0
xgo = 0
bv = 0
Sv = 0
bvo = 0
lvo = 0
Svo = 0
cvo_ = 0
ngo = 0
nvo = 0

# Pylon
bp = 0
cp_ = 0
Sp = 0
npylon = 0

# Fuselage
lf = 0
Df = 0
Smf = 0
lamdaf = 0
Ssm = 0
lnf = 0
lamdanf = 0

# Gondola
lgd = 0
Dgd = 0
lamdagd = 0
Ssm_gd = 0
ln_gd = 0
lamdan_gd = 0
ngd = 0
ngsh = 0

lgsh = 0
Dgsh = 0
lamdagsh = 0
Ssm_gsh = 0
ln_gsh = 0
lamdan_gsh = 0

# Common Data
Dv = 0
Fv = 0
Gvzl = 0
V = 0
number = 0
P0 = 0
H = 0
type = ''
Kint = 0
cxk_light = 0
nlight = 0

# ВЕЛИЧИНЫ ИЗ ГРАФИКОВ
Kx = 0
Kn = 0
Cyamaxprof = 0

# ВЕЛИЧИНЫ ИЗ ТАБЛИЦ
pH = 0  # массовая плотность
aH = 0  # скорость звука
vH = 0 # вязкость воздуха
Gm = 0  # полный запас топлива

dCyamax = 0  # по типу закрылка

# CONST
g = 9.8
p_zero = 1.225  # плотность воздуха на нулевой высоте
v_zero = 1.4607 * math.pow(10, -5)  # коэф кинематической вязкости на нулевой высоте

# ВЫСЧИТАННЫЕ
Mrasch = 0  # число Маха расчетное
cya_rasch = 0 # расчетный КПС
Gpol = 0  # полетный вес самолета
Cyamax = 0
da0_vzl = 0
# Вспомогательная кривая
Re = 0
Vmin = 0

# взлетная кривая


# для интерфейса
permision_for_curve = [False, False, False]
permision_for_export = ['', '', '','']
permissiom_for_polar=['', '', '', '', '']
permission=['','']

# для таблиц
# вспомогательная поляра
linear_size = []
S_proiz=[]
c_lamda_el_for_cxa = []
Sk_el_for_cxa = []
n_el_for_cxa = []
a_list_help=[]
cya_list_help=[]
cxo=0
xtau_el_for_cxa = []
nc_for_cre =[]
nint_for_cre =[]

# взлетная поляра
a_list_up=[]
cya_list_up=[]
Cyamax_vzl =0
a_list_up_scrin=[]
cya_list_up_scrin=[]
Cyamax_vzl_scrin=0
lamdaef_scrin=0

# посадочная кривая
a_list_down=[]
cya_list_down=[]
Cya_max_pos =0
a_list_down_scrin=[]
cya_list_down_scrin=[]
Cya_max_pos_scrin=0
lamdaef_down_scrin=0

# крейсерская
list_Mkr=[]
list_K=[]
# ДИАЛОГОВОЕ ОКНО С РАСЧЕТНЫМИ СХЕМАМИ
class DialogPlan(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.dialog = Ui_Dialog()
        self.dialog.setupUi(self)
        self.setFixedSize(self.size())
        self.setBu()
        self.pixmap = QtGui.QPixmap("images/схема.png")
        self.dialog.label_6.setText('Рисунок 1')
        self.scale = 1
        self.posX, self.posY = 0,0

    def setBu(self):
        self.dialog.buNext.clicked.connect(self.NextPic)
        self.dialog.pushButton.clicked.connect(self.BackPic)
        self.dialog.tabPlan.setTabVisible(0, False)
        self.dialog.tabPlan.setTabVisible(1, False)
        self.dialog.tabPlan.setTabVisible(2, False)
        self.dialog.tabPlan.setTabVisible(3, False)
        self.dialog.tabPlan.setTabVisible(4, False)
        self.dialog.tabPlan.setTabVisible(5, False)
        self.dialog.tabPlan.setCurrentIndex(0)

        self.dialog.label.installEventFilter(self)
        self.dialog.label.setMouseTracking(True)

        self.dialog.label_2.installEventFilter(self)
        self.dialog.label_2.setMouseTracking(True)

        self.dialog.label_3.installEventFilter(self)
        self.dialog.label_3.setMouseTracking(True)

        self.dialog.label_4.installEventFilter(self)
        self.dialog.label_4.setMouseTracking(True)

        self.dialog.label_7.installEventFilter(self)
        self.dialog.label_7.setMouseTracking(True)

        self.dialog.label_5.installEventFilter(self)
        self.dialog.label_5.setMouseTracking(True)

    # def mouseMoveEvent(self, QMouseEvent):
    #     print(QMouseEvent.pos())

    def eventFilter(self, obj, event):
        if (obj is self.dialog.label or self.dialog.label_2 or self.dialog.label_3
            or self.dialog.label_4 or self.dialog.label_7 or self.dialog.label_5):
            if event.type() == QtCore.QEvent.Wheel:
                if event.angleDelta().y() > 0:
                    self.zoomIn(obj)
                else:
                    self.zoomOut(obj)
            if event.type() == QtCore.QEvent.MouseMove:
                self.getMousePosition(event)


        return super(DialogPlan, self).eventFilter(obj, event)

    def getMousePosition(self, event):
        positions = event.pos();
        self.posX , self.posY = positions.x(), positions.y()

    def zoomIn(self, obj):
        self.scale *= 1.1
        self.resize_image(obj)

    def zoomOut(self, obj):
        self.scale /= 1.1
        self.resize_image(obj)

    def resize_image(self, obj):
       size = self.pixmap.size()
       scaled_pixmap = self.pixmap.scaled(int(self.pixmap.width() * self.scale), int(self.pixmap.height() * self.scale), QtCore.Qt.KeepAspectRatio)
       obj.setPixmap(scaled_pixmap)

    def refreshImage(self, obj):
        self.scale = 1
        obj.setPixmap(self.pixmap)

    def NextPic(self):
        index = self.dialog.tabPlan.currentIndex()
        # match index:
        #     case 0:
        #         self.dialog.tabPlan.setCurrentIndex(1)
        #         t = '2'
        if index == 0:
            self.dialog.tabPlan.setCurrentIndex(1)
            self.pixmap = QtGui.QPixmap("images/схема2.png")
            self.refreshImage(self.dialog.label_2)
            t = '2'
        elif index == 1:
            self.dialog.tabPlan.setCurrentIndex(2)
            self.pixmap = QtGui.QPixmap("2.png")
            self.refreshImage(self.dialog.label_3)
            t = '3'
        elif index == 2:
            self.dialog.tabPlan.setCurrentIndex(3)
            self.pixmap = QtGui.QPixmap("1.png")
            self.refreshImage(self.dialog.label_4)
            t = '4'
        elif index == 3:
            self.dialog.tabPlan.setCurrentIndex(4)
            self.pixmap = QtGui.QPixmap("images/носовая часть.png")
            self.refreshImage(self.dialog.label_7)
            t = '5'
        elif index == 4:
            self.dialog.tabPlan.setCurrentIndex(5)
            self.pixmap = QtGui.QPixmap("Схема профиля.png")
            self.refreshImage(self.dialog.label_5)
            t = '6'
        elif index == 5:
            self.dialog.tabPlan.setCurrentIndex(0)
            self.pixmap = QtGui.QPixmap("images/схема.png")
            self.refreshImage(self.dialog.label)
            t = '1'
        self.dialog.label_6.setText('Рисунок '+t)

    def BackPic(self):
        index = self.dialog.tabPlan.currentIndex()

        if index == 0:
            self.dialog.tabPlan.setCurrentIndex(5)
            self.pixmap = QtGui.QPixmap("Схема профиля.png")
            self.refreshImage(self.dialog.label_5)
            t='6'
        elif index == 1:
            self.dialog.tabPlan.setCurrentIndex(0)
            self.pixmap = QtGui.QPixmap("images/схема.png")
            self.refreshImage(self.dialog.label)
            t='1'
        elif index == 2:
            self.dialog.tabPlan.setCurrentIndex(1)
            self.pixmap = QtGui.QPixmap("images/схема2.png")
            self.refreshImage(self.dialog.label_2)
            t='2'
        elif index == 3:
            self.dialog.tabPlan.setCurrentIndex(2)
            self.pixmap = QtGui.QPixmap("2.png")
            self.refreshImage(self.dialog.label_3)
            t='3'
        elif index == 4:
            self.dialog.tabPlan.setCurrentIndex(3)
            self.pixmap = QtGui.QPixmap("1.png")
            self.refreshImage(self.dialog.label_4)
            t='4'
        elif index == 5:
            self.dialog.tabPlan.setCurrentIndex(4)
            self.pixmap = QtGui.QPixmap("images/носовая часть.png")
            self.refreshImage(self.dialog.label_7)
            t='5'

        self.dialog.label_6.setText('Рисунок ' + t)

class Dialog_Help(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.dialog = Ui_Dialog_help()
        self.dialog.setupUi(self)
        self.set()

    def set(self):
        self.dialog.pushButton_2.clicked.connect(self.NextPage)
        self.dialog.pushButton.clicked.connect(self.BackPage)

    def NextPage(self):
        index = self.dialog.tabWidget.currentIndex()
        match index:
            case 0:
                self.dialog.tabWidget.setCurrentIndex(1)
            case 1:
                self.dialog.tabWidget.setCurrentIndex(2)
            case 2:
                self.dialog.tabWidget.setCurrentIndex(3)
            case 3:
                self.dialog.tabWidget.setCurrentIndex(4)
            case 4:
                self.dialog.tabWidget.setCurrentIndex(5)
            case 5:
                self.dialog.tabWidget.setCurrentIndex(6)
            case 6:
                self.dialog.tabWidget.setCurrentIndex(7)
            case 7:
                self.dialog.tabWidget.setCurrentIndex(8)
            case 8:
                self.dialog.tabWidget.setCurrentIndex(9)
            case 9:
                self.dialog.tabWidget.setCurrentIndex(10)
            case 10:
                self.dialog.tabWidget.setCurrentIndex(0)

    def BackPage(self):
            index = self.dialog.tabWidget.currentIndex()
            match index:
                case 0:
                    self.dialog.tabWidget.setCurrentIndex(10)
                case 1:
                    self.dialog.tabWidget.setCurrentIndex(0)
                case 2:
                    self.dialog.tabWidget.setCurrentIndex(1)
                case 3:
                    self.dialog.tabWidget.setCurrentIndex(2)
                case 4:
                    self.dialog.tabWidget.setCurrentIndex(3)
                case 5:
                    self.dialog.tabWidget.setCurrentIndex(4)
                case 6:
                    self.dialog.tabWidget.setCurrentIndex(5)
                case 7:
                    self.dialog.tabWidget.setCurrentIndex(6)
                case 8:
                    self.dialog.tabWidget.setCurrentIndex(7)
                case 9:
                    self.dialog.tabWidget.setCurrentIndex(8)
                case 10:
                    self.dialog.tabWidget.setCurrentIndex(9)

# ДИАЛОГОВОЕ ОКНО С ЗАКРЫЛКАМИ
class DialogMechanism(QtWidgets.QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.dialogM = Ui_Form()
        self.dialogM.setupUi(self)
        self.set()
        self.check()

    def set(self):
        self.dialogM.buClose.clicked.connect(self.ChooseClose)

    def check(self):
        if Flap_type == 'Исходное крыло':
            self.dialogM.rbWing.setChecked(True)
        elif Flap_type =='Простой щиток':
                self.dialogM.rbSimpleBox.setChecked(True)
        elif Flap_type == 'Щиток ЦАГИ':
            self.dialogM.rbBoxCAGI.setChecked(True)
        elif Flap_type == 'Простой закрылок':
            self.dialogM.rbSimpleFlap.setChecked(True)
        elif Flap_type == 'Однощелевой закрылок':
            self.dialogM.rbOneFlap.setChecked(True)
        elif Flap_type == 'Двухщелевой закрылок':
            self.dialogM.rbTwoFlap.setChecked(True)
        elif Flap_type == 'Трехщелевой закрылок':
            self.dialogM.rnThreeFlap.setChecked(True)
        elif Flap_type == 'Закрылок Фаулера':
            self.dialogM.rbFlapFauler.setChecked(True)
        elif Flap_type == 'Двухщелевой закрылок Фаулера':
            self.dialogM.rbTwoFlapFauler.setChecked(True)

    def ChooseClose(self):
        global Flap_type
        if self.dialogM.rbWing.isChecked():
            Flap_type = 'Исходное крыло'
        elif self.dialogM.rbSimpleBox.isChecked():
            Flap_type = 'Простой щиток'
        elif self.dialogM.rbBoxCAGI.isChecked():
            Flap_type = 'Щиток ЦАГИ'
        elif self.dialogM.rbSimpleFlap.isChecked():
            Flap_type = 'Простой закрылок'
        elif self.dialogM.rbOneFlap.isChecked():
            Flap_type = 'Однощелевой закрылок'
        elif self.dialogM.rbTwoFlap.isChecked():
            Flap_type = 'Двухщелевой закрылок'
        elif self.dialogM.rnThreeFlap.isChecked():
            Flap_type = 'Трехщелевой закрылок'
        elif self.dialogM.rbFlapFauler.isChecked():
            Flap_type = 'Закрылок Фаулера'
        elif self.dialogM.rbTwoFlapFauler.isChecked():
            Flap_type = 'Двухщелевой закрылок Фаулера'
        else:
            msg_ = QMessageBox()
            msg_.setIcon(QMessageBox.Critical)
            msg_.setText("Выберите тип механизации.")
            msg_.setWindowTitle("Предупреждение")
            msg_.exec_()

        if Flap_type:
            self.close()


# ОСНОВНОЕ ПРИЛОЖЕНИЕ
class ExampleApp(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.main = Ui_MainWindow()
        self.main.setupUi(self)
        self.show()
        self.set()
        self.connection = sqlite3.connect('bd.db')
        self.cursor = self.connection.cursor()

        self.button = [self.main.buWing, self.main.buFlapMulard, self.main.buTail, self.main.buPylon,
                       self.main.buGondola, self.main.buFuselage, self.main.buCommonData]
        self.button_curve = [self.main.buMkr, self.main.buHelp, self.main.buUp, self.main.buDown, self.main.buCre]
        self.button_polar = [self.main.buHelpPolyr, self.main.buUpPolyr, self.main.buDownPolyr, self.main.buCrePolyr, self.main.buExport]

        self.main.buExport.setEnabled(False)
        self.main.buHelp.setEnabled(False)
        self.main.buUp.setEnabled(False)
        self.main.buDown.setEnabled(False)
        self.main.buCre.setEnabled(False)
        self.main.buHelp.setIcon(QtGui.QIcon('images/x.svg'))
        self.main.groupBox_13.setVisible(False)
        self.main.tabData.setTabEnabled(1, False)
        self.main.tabData.setTabEnabled(2, False)

        # График для Мкр
        self.figure = plt.figure()
        self.canvas = FigureCanvas(self.figure)
        layout = QVBoxLayout()
        layout.addWidget(NavigationToolbar(self.canvas, self))
        layout.addWidget(self.canvas)
        # layout.addWidget(str.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft))

        self.main.plotWidget.setLayout(layout)

        # График для вспомогательной кривой
        self.fhelp = plt.figure()
        self.chelp = FigureCanvas(self.fhelp)
        Vlayout_help = QVBoxLayout()
        Vlayout_help.addWidget(NavigationToolbar(self.chelp, self))
        Vlayout_help.addWidget(self.chelp)
        self.main.helpwidget.setLayout(Vlayout_help)

        # График для взлетной кривой
        self.fUp = plt.figure()
        self.cUp = FigureCanvas(self.fUp)
        Vlayout_Up = QVBoxLayout()
        Vlayout_Up.addWidget(NavigationToolbar(self.cUp, self))
        Vlayout_Up.addWidget(self.cUp)
        self.main.upwidget.setLayout(Vlayout_Up)

        # График для посадочной кривой
        self.fDown = plt.figure()
        self.cDown = FigureCanvas(self.fDown)
        Vlayout_Down = QVBoxLayout()
        Vlayout_Down.addWidget(NavigationToolbar(self.cDown, self))
        Vlayout_Down.addWidget(self.cDown)
        self.main.downwidget.setLayout(Vlayout_Down)

        # График для крейсерской кривой
        self.fCruise = plt.figure()
        self.cCruise = FigureCanvas(self.fCruise)
        Vlayout_Cruise = QVBoxLayout()
        Vlayout_Cruise.addWidget(NavigationToolbar(self.cCruise, self))
        Vlayout_Cruise.addWidget(self.cCruise)
        self.main.cruiswidget.setLayout(Vlayout_Cruise)

        # График для вспомогательной поляры
        self.fP_Help = plt.figure()
        self.cP_Help = FigureCanvas(self.fP_Help)
        Vlayout_P_Help = QVBoxLayout()
        Vlayout_P_Help.addWidget(NavigationToolbar(self.cP_Help, self))
        Vlayout_P_Help.addWidget(self.cP_Help)
        self.main.helpwidget_2.setLayout(Vlayout_P_Help)

        # График для взлетной поляры
        self.fP_Up = plt.figure()
        self.cP_Up = FigureCanvas(self.fP_Up)
        Vlayout_P_Up = QVBoxLayout()
        Vlayout_P_Up.addWidget(NavigationToolbar(self.cP_Up, self))
        Vlayout_P_Up.addWidget(self.cP_Up)
        self.main.upwidget_2.setLayout(Vlayout_P_Up)

        # График для посадочной поляры
        self.fP_Down = plt.figure()
        self.cP_Down = FigureCanvas(self.fP_Down)
        Vlayout_P_Down = QVBoxLayout()
        Vlayout_P_Down.addWidget(NavigationToolbar(self.cP_Down, self))
        Vlayout_P_Down.addWidget(self.cP_Down)
        self.main.downwidget_2.setLayout(Vlayout_P_Down)

        # График для крейсерских и полетных поляр
        self.fP_Cre = plt.figure()
        self.cP_Cre = FigureCanvas(self.fP_Cre)
        Vlayout_P_Cre = QVBoxLayout()
        Vlayout_P_Cre.addWidget(NavigationToolbar(self.cP_Cre, self))
        Vlayout_P_Cre.addWidget(self.cP_Cre)
        self.main.cruiswidget_2.setLayout(Vlayout_P_Cre)

    def set(self):
        # Первая вкладка
        self.main.tabWidget_2.setTabVisible(0, False)
        self.main.tabWidget_2.setTabVisible(1, False)
        self.main.tabWidget_2.setTabVisible(2, False)
        self.main.tabWidget_2.setTabVisible(3, False)
        self.main.tabWidget_2.setTabVisible(4, False)
        self.main.tabWidget_2.setTabVisible(5, False)
        self.main.tabWidget_2.setTabVisible(6, False)
        self.main.buStart.clicked.connect(self.ButtonEnabled)
        self.main.buWing.clicked.connect(self.pressedWing)
        self.main.buFlapMulard.clicked.connect(self.pressedFlapMulard)
        self.main.buTail.clicked.connect(self.pressedTail)
        self.main.buPylon.clicked.connect(self.pressedPylon)
        self.main.buFuselage.clicked.connect(self.pressedFuselage)
        self.main.buGondola.clicked.connect(self.pressedGondola)
        self.main.buCommonData.clicked.connect(self.pressedCommonData)
        self.main.buShowAirPlan.clicked.connect(self.OpenPlan)
        self.main.buFlapChoose.clicked.connect(self.OpenFlapChoose)
        self.main.buShowAirPlan_2.clicked.connect(self.OpenHelpDialog)
        self.main.buCre.clicked.connect(self.pressedPolyr)
        self.main.buExport.clicked.connect(self.pressedExport)
        self.main.buFindK.clicked.connect(self.pressedK)

        # вторая вкладка
        self.main.tabWidget.setTabVisible(0, False)
        self.main.tabWidget.setTabVisible(1, False)
        self.main.tabWidget.setTabVisible(2, False)
        self.main.tabWidget.setTabVisible(3, False)
        self.main.tabWidget.setTabVisible(4, False)

        # третья вкладка
        self.main.tabPolyr.setTabVisible(0,False)
        self.main.tabPolyr.setTabVisible(1,False)
        self.main.tabPolyr.setTabVisible(2,False)
        self.main.tabPolyr.setTabVisible(3,False)

        # Построение вспомогательных кривых
        self.main.buMkr.clicked.connect(self.MakeMkr)
        self.main.buHelp.clicked.connect(self.MakeHelp)
        self.main.buUp.clicked.connect(self.MakeUp)
        self.main.buDown.clicked.connect(self.MakeDown)
        self.main.buCre.clicked.connect(self.MakeCruise)

        # Построение поляр
        self.main.buHelpPolyr.clicked.connect(self.MakeHelpPolyr)
        self.main.buUpPolyr.clicked.connect(self.MakeUpPolyr)
        self.main.buDownPolyr.clicked.connect(self.MakeDownPolyr)
        self.main.buCrePolyr.clicked.connect(self.MakeCruisePolyr)

        # Расчет переменных
        self.main.buCalWing.clicked.connect(self.CalculateWing)
        self.main.buCalFlapMulard.clicked.connect(self.CalculateFlapMulard)
        self.main.buCalTail.clicked.connect(self.CalculateTail)
        self.main.buCalPylon.clicked.connect(self.CalculatePylon)
        self.main.buCalFuselage.clicked.connect(self.CalculateFuselage)
        self.main.buCalGondola.clicked.connect(self.CalculateGondola)
        self.main.buCommonDataSafe.clicked.connect(self.SafeCommonData)

    def ButtonEnabled(self):
        self.main.groupBox_13.setVisible(True)
        self.pressedWing()

    # ОПРЕДЕЛЕНИЕ ПЕРЕМЕННОЙ Кх ПО ГРАФИКУ
    def find_Kx(self, n, x_degree):
        list_x = [0, 10, 20, 30, 40, 50, 55]
        global Kx
        Kx = 0
        if n > 2:
            Kx = 1
            print('Kx = ' + str(Kx))
        else:
            if x_degree in list_x:
                Sql_request = 'SELECT Коэффициент FROM Коэф_Kx ' \
                              'WHERE Стреловидность = %s' % x_degree
                self.cursor.execute(Sql_request)
                Kx = float(self.cursor.fetchone()[0])
                print('Kx = ' + str(Kx))
            else:
                i = 0
                while Kx == 0:
                    if x_degree > list_x[i] and x_degree < list_x[i + 1]:
                        Sql_request = 'SELECT Коэффициент FROM Коэф_Kx ' \
                                      'WHERE Стреловидность = %s' % list_x[i]
                        self.cursor.execute(Sql_request)
                        x_1 = float(self.cursor.fetchone()[0])
                        Sql_request = 'SELECT Коэффициент FROM Коэф_Kx ' \
                                      'WHERE Стреловидность = %s' % list_x[i + 1]
                        self.cursor.execute(Sql_request)
                        x_2 = float(self.cursor.fetchone()[0])
                        middle = (x_1 + x_2) / 2
                        if x_degree % 10 == 5:
                            Kx = middle
                            print('Kx = ' + str(Kx))
                        else:
                            a = list_x[i] + 5
                            if x_degree < a:
                                Kx = (x_1 + middle) / 2
                                print('Kx = ' + str(Kx))
                            else:
                                Kx = (x_2 + middle) / 2
                                print('Kx = ' + str(Kx))
                    else:
                        i = i + 1

    # ОПРЕДЕЛЕНИЕ рН, аН, vH ПО ТАБЛИЦЕ
    def find_with_H(self, H):
        global pH, aH, vH
        pH = 0
        aH = 0
        vH = 0
        h = float('%.1f' % (H / 1000))

        Sql_request = 'SELECT * FROM "Стандартная_атмосфера" ' \
                      'WHERE "Высота(км)" = %s' % str(h)
        self.cursor.execute(Sql_request)
        cor = self.cursor.fetchone()
        # print(cor)
        pHstr = cor[3]
        aHstr = cor[6]
        vHstr = cor[5]

        pH = float(pHstr)
        aH = float(aHstr)
        vH = float(vHstr) * math.pow(10,-5)

        print('pH = ' + str(pH))
        print('aH = ' + str(aH))
        print('vH = ' + str(vH))

    # ОПРЕДЕЛЕНИЕ Gm ПО ТАБЛИЦЕ
    def whatWeight(self, Gvzl):
        if Gvzl <= 20000:
            return '20000'
        if Gvzl <= 40000:
            return '40000'
        if Gvzl <= 80000:
            return '80000'
        if Gvzl <= 120000:
            return '120000'
        else:
            return '120001'

    def findGm(self):
        if type == 'ТРД':
            Sql_request = 'SELECT ТРД FROM "Полный запас топлива Gm" ' \
                          'WHERE "Gвзл, кг" = %s' % self.whatWeight(Gvzl)
            self.cursor.execute(Sql_request)
        elif type == 'ТВД и ПД':
            Sql_request = 'SELECT "ТВД и ПД" FROM "Полный запас топлива Gm" ' \
                          'WHERE "Gвзл, кг" = %s' % self.whatWeight(Gvzl)
            self.cursor.execute(Sql_request)

        return float('%.3f' % (float(self.cursor.fetchone()[0]) * 0.01 * Gvzl))

    # ОПРЕДЕЛЕНИЕ Kn ПО ГРАФИКУ
    def find_Kn(self, n):
        list_n = [1, 1.3, 1.5, 1.8, 2.7, 3]
        index = -1
        if n > 3:
            return -0.01 * n + 0.96
        if n<1:
            return 0.9
        else:
            if n in list_n:
                Sql_request = 'SELECT kn FROM Коэф_Kn ' \
                              'WHERE n = %s' % n
                self.cursor.execute(Sql_request)
                return float(self.cursor.fetchone()[0])
            else:
                i = 0
                while index == -1:
                    if n > list_n[i] and n < list_n[i + 1]:
                        index = i
                    else:
                        i = i + 1
                Sql_request = 'SELECT kn FROM Коэф_Kn ' \
                              'WHERE n = %s' % list_n[index]
                self.cursor.execute(Sql_request)
                first = float(self.cursor.fetchone()[0])
                Sql_request = 'SELECT kn FROM Коэф_Kn ' \
                              'WHERE n = %s' % list_n[index + 1]
                self.cursor.execute(Sql_request)
                second = float(self.cursor.fetchone()[0])
                return (first + second) / 2

    # ОПРЕДЕЛЕНИЕ Cya max профиля ПО ГРАФИКУ

    # ОПРЕДЕНИЕ КОЭФ Cyamax
    def ChooseCymaxprof(self, c_, Re):
        # Выборка
        list_c_ = [0.08, 0.1, 0.12, 0.14, 0.16, 0.18, 0.2]
        list_Re = [1, 2, 3, 4, 5, 6, 7, 8]
        if c_ in list_c_:
            if c_ == 0.08 and 3 <= Re <= 6:
                return 0.1 * Re + 0.8
            elif c_ == 0.1 and 0 <= Re <= 4:
                return 0.13 * Re + 0.85
            elif c_ == 0.12 and 0 <= Re <= 3:
                return 0.17 * Re + 0.83
            elif c_ == 0.14 and 0 <= Re <= 3:
                return 0.17 * Re + 0.83
            if Re in list_Re:
                Sql_request = 'SELECT Cyamaxprof FROM Коэф_Cyamaxprof ' \
                              'WHERE Re = %s AND Толщина= %s' % (Re, c_)
                self.cursor.execute(Sql_request)
                return float(self.cursor.fetchone()[0])
            elif Re > 8:
                Sql_request = 'SELECT Cyamaxprof FROM Коэф_Cyamaxprof ' \
                              'WHERE Re = %s AND Толщина= %s' % (8, c_)
                self.cursor.execute(Sql_request)
                return float(self.cursor.fetchone()[0])
            else:
                i = 0
                index = -1
                while index == -1:
                    if list_Re[i] < Re < list_Re[i + 1]:
                        index = i
                    else:
                        i = i + 1
                Sql_request = 'SELECT Cyamaxprof FROM Коэф_Cyamaxprof ' \
                              'WHERE Re = %s AND Толщина= %s' % (list_Re[index], c_)
                self.cursor.execute(Sql_request)
                first = float(self.cursor.fetchone()[0])
                # print(first)
                Sql_request = 'SELECT Cyamaxprof FROM Коэф_Cyamaxprof ' \
                              'WHERE Re = %s AND Толщина= %s' % (list_Re[index + 1], c_)
                self.cursor.execute(Sql_request)
                second = float(self.cursor.fetchone()[0])
                # print(second)
                return float('%.3f' % ((first + second) / 2))

    def find_Cyamaxprof(self):
        global Cyamaxprof, Re, Vmin
        v0 = 0.000014607  # коэф кинематической вязкости на высоте 0
        # определение числа Рейнольдса
        Vmin = 3.5 * math.sqrt(Gpol / S)
        Re_full = (Vmin * b) / v0
        Re = float('%.1f' % (Re_full / 10 ** 6))
        print('Re = ' + str(Re))
        print('c_ = ' + str(c_))

        # Вызов функции
        list_c_ = [0.08, 0.1, 0.12, 0.14, 0.16, 0.18, 0.2]
        if c_ in list_c_:
            Cyamaxprof = self.ChooseCymaxprof(c_, Re)
            print('Cyamaxprof = ' + str(Cyamaxprof))
        elif c_ < 0.08:
            Cyamaxprof = self.ChooseCymaxprof(0.08, Re)
        else:
            i = 0
            index = -1
            while index == -1:
                if list_c_[i] < c_ < list_c_[i + 1]:
                    index = i
                else:
                    i = i + 1
            first = self.ChooseCymaxprof(list_c_[index], Re)
            second = self.ChooseCymaxprof(list_c_[index + 1], Re)
            Cyamaxprof = float('%.3f' % ((first + second) / 2))
            print('Cyamaxprof = ' + str(Cyamaxprof))

    # ВЫЗОВ ФУНКЦИИ .ОПРЕДЕЛЕЛЕНИЯ ПРИРАЩЕНИЯ УГЛА. взависимости от относ хорды закрылков
    def call_da0(self, delta):
        list_bzak_ = [0.1, 0.2, 0.3]
        if bzak_ in list_bzak_:
            return (-1) * self.find_da0(bzak_, delta)
        elif bzak_ > 0.3:
            return (-1) * self.find_da0(0.3, delta)
        else:
            i = 0
            index = -1
            while index == -1:
                if list_bzak_[i] < bzak_ < list_bzak_[i + 1]:
                    index = i
                    first = self.find_da0(list_bzak_[index], delta)
                    second = self.find_da0(list_bzak_[index + 1], delta)
                    if (bzak_ * 100) % 10 == 5:
                        return (-1) * float('%.3f' % ((first + second) / 2))
                    elif (bzak_ * 100) % 10 < 5:
                        middle_a0_vzl = float('%.3f' % ((first + second) / 2))
                        return (-1) * float('%.3f' % ((first + middle_a0_vzl) / 2))
                    elif (bzak_ * 100) % 10 > 5:
                        middle_a0_vzl = float('%.3f' % ((first + second) / 2))
                        return (-1) * float('%.3f' % ((second + middle_a0_vzl) / 2))
                else:
                    i = i + 1

    # ОПРЕДЕЛЕНИЕ ПРИРАЩЕНИЯ УГЛА (взлетного и посадочного)
    def find_da0(self, bzak_, delta):
        list_b_ = [0.1, 0.2, 0.3]
        list_delta = [0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1, 1.1, 1.2, 1.3]
        if bzak_ in list_b_:
            if delta in list_delta:
                    Sql_request = 'SELECT delta_a0 FROM "Приращение взлетного угла" ' \
                                  'WHERE b_zak = %s AND бvzl_pos= %s' % (bzak_, delta)
                    self.cursor.execute(Sql_request)
                    return float(self.cursor.fetchone()[0])
            elif delta > 1.3:
                Sql_request = 'SELECT delta_a0 FROM "Приращение взлетного угла" ' \
                              'WHERE b_zak = %s AND бvzl_pos= %s' % (bzak_, 1.3)
                self.cursor.execute(Sql_request)
                return float(self.cursor.fetchone()[0])
            else:
                i = 0
                index = -1
                while index == -1:
                    if list_delta[i] < delta < list_delta[i + 1]:
                        index = i
                    else:
                        i = i + 1
                Sql_request = 'SELECT delta_a0 FROM "Приращение взлетного угла" ' \
                              'WHERE b_zak = %s AND бvzl_pos= %s' % (bzak_, list_delta[index])
                self.cursor.execute(Sql_request)
                first = float(self.cursor.fetchone()[0])
                # print(first)
                Sql_request = 'SELECT delta_a0 FROM "Приращение взлетного угла" ' \
                              'WHERE b_zak = %s AND бvzl_pos= %s' % (bzak_, list_delta[index + 1])
                self.cursor.execute(Sql_request)
                second = float(self.cursor.fetchone()[0])
                # print(second)
                return float('%.3f' % (self.cal_system(list_delta[index], first , list_delta[index + 1], second, delta)))

    # ОПРЕДЕЛЕНИЕ КОЭФФИЦИЕНТА СОПРОТИВЛЕНИЯ ТРЕНИЯ ПЛОСКОЙ ПЛАСТИНЫ 2Cf
    def find_2cf(self, xt, Re):
        list_Re = [1, 2, 5, 10, 20, 50, 100, 200, 500]
        list_xt = [0, 0.1, 0.2, 0.3, 0.4, 0.5, 1]
        two_cf = 0
        const_v = 0

        if xt == 0.1:
            xt = 0
            if Re >= 11:
                const_v = 0.00025
            else:
                const_v = 0.0004
        # для необозначенных на графике точек перехода
        elif 0.5 < xt < 1:
            multip = (xt - 0.5) * 10
            xt = 0.5
            f = False
            i = 0
            while not f:
                if list_Re[i] <= Re < list_Re[i + 1]:
                    f = True
                    const_v = multip * (0.0007 - (0.000039 * i))
                elif Re == 500:
                    f = True
                    const_v = multip * (0.0007 - (0.000039 * 9))
                elif Re < 1:
                    f = True
                    const_v = multip * (0.0007 - (0.000039 * 0))
                else:
                    i = i + 1

        if xt in list_xt:
            try:
                two_cf = self.request_Kcf(xt, Re)
            except TypeError:
                flag = False
                i = 0
                while not flag:
                    if list_Re[i] < Re < list_Re[i + 1]:
                        flag = True
                        y1 = self.request_Kcf(xt, list_Re[i])
                        y2 = self.request_Kcf(xt, list_Re[i + 1])
                        two_cf = self.cal_system(list_Re[i], y1, list_Re[i + 1], y2, Re)
                    elif Re > 500:
                        two_cf = self.request_Kcf(xt, 500)
                        flag = True
                    elif Re < 1:
                        two_cf = self.request_Kcf(xt, 1)
                        flag = True
                    else:
                        i = i + 1
        else:
            if xt > 1:
                return self.find_2cf(1, Re)
            # для не целых точек перехода
            else:
                flag = False
                i = 0
                while not flag:
                    if list_xt[i] < xt < list_xt[i + 1]:
                        flag = True
                        if xt - list_xt[i] >= 0.03 and list_xt[i + 1] - xt >= 0.03:
                            v1 = self.find_2cf(list_xt[i], Re)
                            v2 = self.find_2cf(list_xt[i + 1], Re)
                            two_cf = (v1 + v2) / 2
                        elif xt - list_xt[i] < 0.03:
                            two_cf = self.find_2cf(list_xt[i], Re)
                        else:
                            two_cf = self.find_2cf(list_xt[i + 1], Re)
                    else:
                        i = i + 1

        # print(two_cf)
        two_cf = float('%.5f' % (two_cf - const_v))
        return two_cf

    # Поиск коэффициентов системы функции и расчет значения
    def cal_system(self, x1, y1, x2, y2, x):
        k = (y1 - y2) / (x1 - x2)
        b_koef = y1 - k * x1
        return float('%.5f' % (k * x + b_koef))

    # SQL запрос для коэффицента 2Cf
    def request_Kcf(self, xt, Re):
        Sql_request = 'SELECT "2cf" FROM Коэф_cf ' \
                      'WHERE "Точка перехода" = %s AND "Re * 10^6"= %s' % (xt, Re)
        self.cursor.execute(Sql_request)
        return float(self.cursor.fetchone()[0])

    # ОПРЕДЕЛЕНИЕ КОЭФФИЦИЕНТА РЕЖИМА ТЕЧЕЧЕНИЯ В ПОГРАНИЧНОМ СЛОЕ ДЛЯ КРЫЛА
    def find_nc_wing(self, xt, c_):
        # list_c_ = [0, 0.01, 0.02, 0.03, 0.04, 0.05, 0.06, 0.07, 0.08, 0.09, 0.1, 0.11, 0.12, 0.13, 0.14, 0.15]
        list_c_ = [float(i) for i in np.arange(0, 0.16, 0.01)]
        cur = self.request_nc_wing(c_)
        if cur is None:
            if c_ > 0.15:
                cur = self.request_nc_wing(0.15)
            elif c_ < 0.01:
                cur = self.request_nc_wing(0)
            else:
                flag = False
                i = 0
                while not flag:
                    if list_c_[i] < c_ < list_c_[i + 1]:
                        y1_cur = self.request_nc_wing(list_c_[i])
                        y2_cur = self.request_nc_wing(list_c_[i + 1])
                        y1 = y1_cur[1] - xt * 10 * y1_cur[2]
                        y2 = y2_cur[1] - xt * 10 * y2_cur[2]
                        return self.cal_system(list_c_[i], y1, list_c_[i + 1], y2, c_)
                    else:
                        i = i + 1

        nc = float('%.4f' % (cur[1] - xt * 10 * cur[2]))
        if nc < 1:
            nc = 1

        return nc

    # SQL запрос для коэффицента nc для крыла
    def request_nc_wing(self, c_):
        Sql_request = 'SELECT * FROM Коэф_nc_крыло WHERE "c_"= %s' % c_
        self.cursor.execute(Sql_request)
        return self.cursor.fetchone()

    # ОПРЕДЕЛЕНИЕ КОЭФФИЦИЕНТА РЕЖИМА ТЕЧЕЧЕНИЯ В ПОГРАНИЧНОМ СЛОЕ ДЛЯ ТЕЛ ВРАЩЕНИЯ
    def find_nc_body_rotate(self, lamda):
        list_lamda = [2.25, 2.6, 3.2, 4, 4.8, 6, 8, 10, 12, 14]
        if lamda in list_lamda:
            return self.request_nc_body_rotate(lamda)
        elif lamda > 14:
            return self.request_nc_body_rotate(14)
        elif lamda < 2.25:
            return self.request_nc_body_rotate(2.25)
        else:
            flag = False
            i = 0
            while not flag:
                if list_lamda[i] < lamda < list_lamda[i + 1]:
                    y1 = self.request_nc_body_rotate(list_lamda[i])
                    y2 = self.request_nc_body_rotate(list_lamda[i + 1])
                    return self.cal_system(list_lamda[i], y1, list_lamda[i + 1], y2, lamda)
                else:
                    i = i + 1

    # SQL запрос для коэффицента nc для крыла
    def request_nc_body_rotate(self, lamda):
        Sql_request = 'SELECT nc FROM Коэф_nc_тело_вращения WHERE "lamda"= %s' % lamda
        self.cursor.execute(Sql_request)
        return self.cursor.fetchone()[0]

    def iconbutton(self, bufill, buttons):
        for bu in buttons:
            bu.setIcon(QtGui.QIcon('images/caret-right.svg'))
        if S == 0 and b == 0:
            self.main.buFlapMulard.setIcon(QtGui.QIcon('images/x.svg'))
        bufill.setIcon(QtGui.QIcon('images/caret-right-fill.svg'))

    # ОПРЕДЕЛЕНЕ КОЭФФИЦИЕНТА Кинт
    def request_Kint(self, type):
        Sql_request = 'SELECT "Кинт" FROM "Коэф_Кинт" WHERE "Форма" = %s' % str('"' + type + '"')
        self.cursor.execute(Sql_request)
        return self.cursor.fetchone()[0]

    # ОПРЕДЕЛЕНИЕ КОЭФФИЦИЕНТА cxk для фонаря кабины пилотов
    def request_cxk(self, type):
        Sql_request = 'SELECT "cxk" FROM "Коэф_cxk" WHERE "Фонарь" = %s' % str('"' + type + '"')
        self.cursor.execute(Sql_request)
        return self.cursor.fetchone()[0]

    # ОПРЕДЕЛЕНИЕ ПОПРАВКИ
    def call_delta(self, lamda, n):
        list_lamda = [5, 7, 10]
        delta = 0

        if lamda in list_lamda:
            delta = self.find_delta(lamda, n)
        elif lamda > 10:
            delta = self.find_delta(10, n)
        elif lamda < 5:
            delta = self.find_delta(5, n)
        elif 5 < lamda < 7:
            first = self.find_delta(5, n)
            second = self.find_delta(7, n)
            part = (second - first) / 20
            print(lamda % 5)
            delta = (lamda % 5) * part * 10 + first
        elif 7 < lamda < 10:
            first = self.find_delta(7, n)
            second = self.find_delta(10, n)
            part = (second - first) / 30
            delta = (lamda % 7) * part * 10 + first

        return float('%.4f' % delta)

    def find_delta(self, lamda, n):
        list_n = [float(i) for i in np.arange(1.25, 5.25, 0.25)]
        delta = 0
        try:
            delta = self.request_delta(lamda, n)
        except:
            if n < 1.25:
                delta = self.request_delta(lamda, 1.25)
            elif n > 5:
                y1 = self.call_delta(lamda, 4.5)
                y2 = self.call_delta(lamda, 5)
                print(y1)
                print(y2)
                return self.cal_system(4.5, y1, 5, y2, n)
            else:
                flag = False
                i = 0
                while not flag:
                    if list_n[i] < n < list_n[i + 1]:
                        y1 = self.request_delta(lamda, list_n[i])
                        y2 = self.request_delta(lamda, list_n[i + 1])
                        return self.cal_system(list_n[i], y1, list_n[i + 1], y2, n)
                    else:
                        i += 1
        return float(delta)

    def request_delta(self, lamda, n):
        Sql_request = 'SELECT Дельта FROM Поправка ' \
                      'WHERE "Удлинение"= %s AND "Сужение"= %s' % (lamda, n)
        self.cursor.execute(Sql_request)
        return self.cursor.fetchone()[0]

    # ОПРЕДЕЛЕНИЕ КОЭФФИЦИЕНТА ПРИРАЩЕНИЯ ОТ ВЫПУЩЕННЫХ ЗАКРЫЛКОВ
    def find_dCxo_zak(self, bzak, delta):
        list_delta = [float('%.1f' % i) for i in np.arange(0, 0.9, 0.1)] # МОЖЕТ БЫТЬ 1 ???

        if (bzak == 0.3 or bzak == 0.1) and delta in list_delta:
            return self.request_dcxo(bzak, delta)
        elif delta > 0.8:
            y1 = self.find_dCxo_zak(bzak, 0.7)
            y2 = self.find_dCxo_zak(bzak, 0.8)
            return self.cal_system(0.7, y1, 0.8, y2, delta)
        elif bzak == 0.3 or bzak == 0.1:
            flag = False
            i = 0
            while not flag:
                if list_delta[i] < delta < list_delta[i + 1]:
                    y1 = self.request_dcxo(bzak, list_delta[i])
                    y2 = self.request_dcxo(bzak, list_delta[i + 1])
                    return self.cal_system(list_delta[i], y1, list_delta[i + 1], y2, delta)
                else:
                    i += 1
        else:
            first = self.find_dCxo_zak(0.1, delta)
            second = self.find_dCxo_zak(0.3, delta)
            part = (second - first) / 20
            dcxo_zak = (bzak - 0.1) * part * 100 + first
            return float('%.4f' % dcxo_zak)

    def request_dcxo(self, b, d):
        Sql_request = 'SELECT dcxo_zak FROM dCxo_zak ' \
                      'WHERE b_zak= %s AND delta= %s' % (b, d)
        self.cursor.execute(Sql_request)
        return self.cursor.fetchone()[0]


    # ОПРЕДЕЛЕНИЕ КОЭФ nM ТЕЛО ВРАЩЕНИЯ
    def call_nM_body_rotation(self, la_nch, M):
        list_la_nch = [1,2,3,4,5]
        nM = 0

        if la_nch in list_la_nch:
            nM = self.find_nM_body_rotation(la_nch, M)
        elif la_nch > 5:
            nM = self.find_nM_body_rotation(6, M)
        elif la_nch < 1:
            nM = self.find_nM_body_rotation(1, M)
        else:
            flag = False
            i = 0
            while not flag:
                if list_la_nch[i] < la_nch < list_la_nch[i + 1]:
                    first = self.find_nM_body_rotation(list_la_nch[i], M)
                    second = self.find_nM_body_rotation(list_la_nch[i+1], M)
                    #print(first)
                    #print(second)
                    part =  (first - second) / 10
                    nM = first - (la_nch - list_la_nch[i]) * part *10
                    flag = True
                else:
                    i += 1

        return float('%.4f' % nM)

    def find_nM_body_rotation(self, la_nch, M):
        list_M = [float('%.1f' % i) for i in np.arange(0.2, 1, 0.1)]
        nM = 0
        try:
            nM = self.request_nM_body_rotation(la_nch, M)
        except:
            if M < 0.2:
                nM = 1
            elif M > 0.9:
                y1 = self.call_nM_body_rotation(la_nch, 0.8)
                y2 = self.call_nM_body_rotation(la_nch, 0.9)
                return self.cal_system(0.8, y1, 0.9, y2, M)
            else:
                flag = False
                i = 0
                while not flag:
                    if list_M[i] < M < list_M[i + 1]:
                        y1 = self.request_nM_body_rotation(la_nch, list_M[i])
                        y2 = self.request_nM_body_rotation(la_nch, list_M[i + 1])
                        return self.cal_system(list_M[i], y1, list_M[i + 1], y2, M)
                    else:
                        i += 1
        return float(nM)

    def request_nM_body_rotation(self, la_nch, M):
        Sql_request = 'SELECT nM FROM Коэф_nM_тело_вращения ' \
                      'WHERE lamda_nch= %s AND M= %s' % (la_nch, M)
        self.cursor.execute(Sql_request)
        return self.cursor.fetchone()[0]

        # ОПРЕДЕЛЕНИЕ КОЭФ nM КРЫЛО
    def call_nM_wing(self, c_, M):
        list_c_ = [0,4,8,12,20]
        nM = 0

        if c_ in list_c_:
            nM = self.find_nM_wing(c_, M)
        elif c_ > 20:
            nM = self.find_nM_wing(20, M)
        else:
            flag = False
            i = 0
            while not flag:
                if list_c_[i] < c_ < list_c_[i + 1]:
                    first = self.find_nM_wing(list_c_[i], M)
                    second = self.find_nM_wing(list_c_[i + 1], M)
                    # print(first)
                    # print(second)
                    r = list_c_[i + 1] - list_c_[i]
                    part = (second - first) / (r*10)
                    nM = first + (c_ - list_c_[i]) * part * 10
                    flag = True
                else:
                    i += 1

        return float('%.4f' % nM)

    def find_nM_wing(self, c_, M):
        list_M = [float('%.1f' % i) for i in np.arange(0.2, 1, 0.1)]
        nM = 0
        try:
            nM = self.request_nM_wing(c_, M)
        except:
            if M < 0.2:
                nM = 1
            elif M > 0.9:
                y1 = self.call_nM_wing(c_, 0.8)
                y2 = self.call_nM_wing(c_, 0.9)
                return self.cal_system(0.8, y1, 0.9, y2, M)
            else:
                flag = False
                i = 0
                while not flag:
                    if list_M[i] < M < list_M[i + 1]:
                        y1 = self.request_nM_wing(c_, list_M[i])
                        y2 = self.request_nM_wing(c_, list_M[i + 1])
                        return self.cal_system(list_M[i], y1, list_M[i + 1], y2, M)
                    else:
                        i += 1
        return float(nM)

    def request_nM_wing(self, c_, M):
        Sql_request = 'SELECT nM FROM Коэф_nM_крыло ' \
                      'WHERE c_= %s AND M= %s' % (c_, M)
        self.cursor.execute(Sql_request)
        return self.cursor.fetchone()[0]

    # ОТКРЫТИЕ ВКЛАДОК ПО КНОПКАМ
    def pressedWing(self):
        self.main.tabWidget_2.setCurrentIndex(0)

        # v = self.find_2cf(0.4, 500)
        # print('2cf = '+str(v))
        # nc_wing = self.find_nc_wing(0.5, 0.16)
        # print('nc_wing = ' + str(nc_wing))
        # nc_br = self.find_nc_body_rotate(8)
        # print('nc_br = ' + str(nc_br))
        # print('delta = ' + str(self.call_delta(2.2, 13)))
        # print('dCxo_zak = ' + str(self.find_dCxo_zak(0.2, 0.68)))
        # print('nM тело = ' + str(self.call_nM_body_rotation(2.7, 0.95)))
        # print('nM крыло = ' + str(self.call_nM_body_rotation(3.116, 0.7)))

        self.iconbutton(self.main.buWing, self.button)

    def pressedFlapMulard(self):
        if b != 0 and S != 0:
            self.main.tabWidget_2.setCurrentIndex(1)
            self.iconbutton(self.main.buFlapMulard, self.button)
        else:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setText("Сначала рассчитайте параметры крыла.")
            msg.setWindowTitle("Предупреждение")
            msg.exec_()

    def pressedTail(self):
        self.main.tabWidget_2.setCurrentIndex(2)
        self.iconbutton(self.main.buTail, self.button)

    def pressedPylon(self):
        self.main.tabWidget_2.setCurrentIndex(3)
        self.iconbutton(self.main.buPylon, self.button)

    def pressedFuselage(self):
        self.main.tabWidget_2.setCurrentIndex(4)
        self.iconbutton(self.main.buFuselage, self.button)

    def pressedGondola(self):
        self.main.tabWidget_2.setCurrentIndex(5)
        self.iconbutton(self.main.buGondola, self.button)

    def pressedCommonData(self):
        self.main.tabWidget_2.setCurrentIndex(6)
        self.iconbutton(self.main.buCommonData, self.button)

    def pressedPolyr(self):
        self.main.tabData.setTabEnabled(2, True)
        # self.MakeHelpPolyr()

    def pressedExport(self):
        self.iconbutton(self.main.buExport, self.button_polar)
        try:
            file_name = QtWidgets.QFileDialog.getSaveFileName(self, 'Сохранение файла', '', filter='*.xlsx')
            if file_name:
                    self.ExportImage(str(file_name[0]))
        except PermissionError:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Убедитесь, что файл, который Вы редактируете, закрыт.")
            msg.setWindowTitle("Ошибка")
            msg.exec_()
        except:
            self.MakeCruisePolyr()

    def OpenPlan(self):
        dlg = DialogPlan(self)
        dlg.show()

    def OpenFlapChoose(self):
        dialog = DialogMechanism(self)
        dialog.show()

    def OpenHelpDialog(self):
        dialog_help = Dialog_Help(self)
        dialog_help.show()

    # РАСЧЕТ ИСХОДНЫХ ДАННЫХ КРЫЛА
    def CalculateWing(self):

        try:
            # Хорда средняя
            global l, S, b
            l = float(self.main.ed_l.text().replace(',', '.'))
            S = float(self.main.ed_S.text().replace(',', '.'))
            b = float('%.3f' % (S / l))
            self.main.ed_b.setText(str(b))
            if b != 0 and S != 0:
                self.main.buFlapMulard.setCursor(QCursor(QtCore.Qt.CursorShape.PointingHandCursor))
                self.main.buFlapMulard.setStyleSheet("background-color: white;")

            # Сужение
            global b0, bk, n
            b0 = float(self.main.ed_b0.text().replace(',', '.'))
            bk = float(self.main.ed_bk.text().replace(',', '.'))
            n = float('%.2f' % (b0 / bk))
            self.main.ed_n.setText(str(n))

            # Относительная толщина профиля
            global c_
            c_ = float(self.main.dsp_c_.text().replace(',', '.'))

            # Относительная координата максимальной толщины
            global xc_
            xc_ = float(self.main.dsp_xc_.text().replace(',', '.'))

            # Относительная кривизна профиля
            global f_
            f_ = float(self.main.dsp_f_.text().replace(',', '.'))

            # Угол атаки нулевой подъемной силы
            global a0
            expression = (-1) * 0.9 * f_
            a0 = float('%.3f' % expression)
            self.main.ed_a0.setText(str(a0))

            # Относительная координата фокуса профиля
            global xf_
            xf_ = float(self.main.dsp_xf_.text().replace(',', '.'))

            # Удлинение геометрическое
            global lamda
            lamda = float('%.3f' % (l ** 2 / S))
            self.main.ed_lamda.setText(str(lamda))

            # Относительная площадь, занятая фюзеляжем
            global Sf, Sf_
            Sf = float(self.main.ed_Sf.text().replace(',', '.'))
            Sf_ = float('%.3f' % (Sf / S))
            self.main.ed_Sf_.setText(str(Sf_))

            # Относительная площадь, занятая гондолами двигателей
            global Sgd, Sgd_
            Sgd = float(self.main.ed_Sgd.text().replace(',', '.'))
            Sgd_ = float('%.3f' % (Sgd / S))
            self.main.ed_Sgd_.setText(str(Sgd_))

            # Относительная площадь, занятая гондолами шасси
            global Sgsh, Sgsh_
            Sgsh = float(self.main.ed_Sgsh.text().replace(',', '.'))
            Sgsh_ = float('%.3f' % (Sgsh / S))
            self.main.ed_Sgsh_.setText(str(Sgsh_))

            # Относительная площадь, не обтекаемая потоком
            global S_
            S_ = round(Sf_ + Sgd_ + Sgsh_, 3)
            self.main.ed_S_.setText(str(S_))

            # Удлинение эффективное
            global lamdaef, Kx, x_degree
            x_degree = float(self.main.ed_x_deg.text().replace(',', '.'))
            self.find_Kx(n, x_degree)
            expression = (lamda * Kx) / (1 + S_)
            lamdaef = float('%.3f' % expression)
            self.main.ed_lamdaef.setText(str(lamdaef))

            # Производная коэф подъемной силы по углу атаки
            global caya, xc_degree
            xc_degree = float(self.main.ed_xc_deg.text().replace(',', '.'))
            if x_degree > 90 or xc_degree > 90:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Critical)
                msg.setText("Стреловидность не может быть больше 90 градусов.")
                msg.setWindowTitle("Ошибка")
                msg.exec_()
            else:
                xc_degree = float(self.main.ed_xc_deg.text().replace(',', '.')) * math.pi / 180
                expression = (2 * math.pi * lamdaef * math.cos(xc_degree)) / (
                        57.3 * (lamdaef + 2 * math.cos(xc_degree)))
                caya = float('%.3f' % expression)
                self.main.ed_caya.setText(str(caya))

                # Относительная площадь, обдуваемая винтами
                global Sobd, Sobd_
                Sobd = float(self.main.ed_Sobd.text().replace(',', '.'))
                Sobd_ = float('%.3f' % (Sobd / S))
                self.main.ed_Sobd_.setText(str(Sobd_))

                # Относительная координата точки перехода ЛПС в ТПС
                global xtau_
                if xc_ == 0:
                    expression = xc_ * (1 - Sobd_)
                else:
                    expression = 0
                xtau_ = float('%.3f' % expression)
                self.main.ed_xtau_.setText(
                    str(xtau_))  # У РЕБЯТ РАВНО 0 !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!

                # Коэффициент момента профиля
                global cmo
                cmo = float('%.3f' % ((-1) * 0.005 * math.pi * f_))
                self.main.ed_cmo.setText(str(cmo))

                # Определение переменных
                global h
                h = float(self.main.ed_h.text().replace(',', '.'))

            self.iconbutton(self.main.buWing, self.button)
            permision_for_curve[0] = True
        except:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Неправильный ввод данных.")
            msg.setWindowTitle("Ошибка")
            msg.setDetailedText("Вводить можно только цифры. Числа должны быть положительными.")
            msg.exec_()

        if permision_for_curve[0] and permision_for_curve[1] and permision_for_curve[2]:
            self.main.tabData.setTabEnabled(1, True)

    # РАСЧЕТ ЗАКРЫЛОК И ПРЕДКРЫЛОК
    def CalculateFlapMulard(self):

        try:
            # Определение переменных
            global xshzak
            xshzak = float(self.main.ed_xshzak.text().replace(',', '.'))

            # Относительная хорда закрылки
            global bzak_
            bzak_ = float(self.main.dsb_bzak_.text().replace(',', '.'))

            # Относ площадь крыла, обслуживаемая закрылками
            global Sobzak, Sobzak_
            Sobzak = float(self.main.ed_Sobzak.text().replace(',', '.'))
            Sobzak_ = float('%.3f' % (Sobzak / S))
            self.main.ed_Sobzak_.setText(str(Sobzak_))

            # Хорда средняя крыла с выпущенными закрылком
            global bsrzak, lzak
            lzak = float(self.main.ed_lzak.text().replace(',', '.'))
            bsrzak = float('%.3f' % (Sobzak / lzak))
            self.main.ed_bsrzak.setText(str(bsrzak))

            # Расстояние от края закрылка до земли при взлете
            global deltavzl, hvzl, deltapos
            deltavzl = float(self.main.sb_deltavzl.text().replace(',', '.'))* math.pi / 180
            deltapos = float(self.main.sb_deltapos.text().replace(',', '.')) * math.pi / 180
            hvzl = float('%.3f' % (h - math.sin(deltavzl) * bzak_*b))  # !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
            self.main.ed_hvzl.setText(str(hvzl))

            # Расстояние от края закрылка до земли при посадке
            global hpos
            hpos = float('%.3f' % (h - math.sin(deltapos) * bzak_*b))  # !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
            self.main.ed_hpos.setText(str(hpos))

            # Относ площадь крыла, обслуживаемая предкрылком
            global Sobpr, Sobpr_
            Sobpr = float(self.main.ed_Sobpr.text().replace(',', '.'))
            Sobpr_ = float('%.3f' % (Sobpr / S))
            self.main.ed_Sobpr_.setText(str(Sobpr_))

            permision_for_curve[1] = True

        except ValueError:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Неправильный ввод данных.")
            msg.setWindowTitle("Ошибка")
            msg.setDetailedText("Вводить можно только цифры. Числа должны быть положительными.")
            msg.exec_()

        if permision_for_curve[0] and permision_for_curve[1] and permision_for_curve[2]:
            self.main.tabData.setTabEnabled(1, True)

    # РАСЧЕТ ГОРИЗОНТ И ВЕРТИКАЛ ОПЕРЕНИЯ
    def CalculateTail(self):

        try:
            # Определение переменных
            global bgo, xgo, bv, Sv, bvo, lvo, Svo
            bgo = float(self.main.ed_bgo.text().replace(',', '.'))
            xgo = float(self.main.ed_xgo.text().replace(',', '.'))
            bv = float(self.main.ed_bv.text().replace(',', '.'))
            Sv = float(self.main.ed_Sv.text().replace(',', '.'))
            bvo = float(self.main.ed_bvo.text().replace(',', '.'))
            lvo = float(self.main.ed_lvo.text().replace(',', '.'))
            Svo = float(self.main.ed_Svo.text().replace(',', '.'))

            # Относительная толщина горизонт и вертикал оперения
            global cgo_, cvo_
            cvo_ = cgo_ = float('%.3f' % (c_ - 0.02))
            self.main.ed_cgo_.setText(str(cgo_))
            self.main.ed_cvo_.setText(str(cvo_))

            # Удлинение гор оперения
            global lgo, Sgo, lamdago
            lgo = float(self.main.ed_lgo.text().replace(',', '.'))
            Sgo = float(self.main.ed_Sgo.text().replace(',', '.'))
            lamdago = float('%.3f' % (lgo ** 2 / Sgo))
            self.main.ed_lamdago.setText(str(lamdago))

            global nvo, ngo
            ngo = float(self.main.ed_ngo.text().replace(',', '.'))
            nvo = float(self.main.ed_nvo.text().replace(',', '.'))

        except ValueError:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Неправильный ввод данных.")
            msg.setWindowTitle("Ошибка")
            msg.setDetailedText("Вводить можно только цифры. Числа должны быть положительными.")
            msg.exec_()

        if permision_for_curve[0] and permision_for_curve[1] and permision_for_curve[2]:
            self.main.tabData.setTabEnabled(1, True)

    # ОПРЕДЕЛЕНИЕ ПЕРЕМЕННЫХ ПИЛОНА
    def CalculatePylon(self):

        try:
            global bp, cp_, Sp, npylon
            bp = float(self.main.ed_bp.text().replace(',', '.'))
            cp_ = float(self.main.ed_cp_.text().replace(',', '.'))
            Sp = float(self.main.ed_Sp.text().replace(',', '.'))
            npylon = float(self.main.ed_np.text().replace(',', '.'))

        except ValueError:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Неправильный ввод данных.")
            msg.setWindowTitle("Ошибка")
            msg.setDetailedText("Вводить можно только цифры. Числа должны быть положительными.")
            msg.exec_()

        if permision_for_curve[0] and permision_for_curve[1] and permision_for_curve[2]:
            self.main.tabData.setTabEnabled(1, True)


    # РАСЧЕТ ДАННЫХ ФЮЗЕЛЯЖА
    def CalculateFuselage(self):

        try:
            # Диаметр миделя
            global Df, Smf
            Smf = float(self.main.ed_Smf.text().replace(',', '.'))
            Df = float('%.3f' % (math.sqrt(4 * Smf / math.pi)))
            self.main.ed_Df.setText(str(Df))

            # Удлинение
            global lamdaf, lf
            lf = float(self.main.ed_lf.text().replace(',', '.'))
            lamdaf = float('%.3f' % (lf / Df))
            self.main.ed_lamdaf.setText(str(lamdaf))

            # Смоченная поверхность
            global Ssm
            Ssm = float('%.3f' % (2.5 * lf * Df))
            self.main.ed_Ssm.setText(str(Ssm))

            # Удлинение носовой части
            global lnf, lamdanf
            lnf = float(self.main.ed_lnf.text().replace(',', '.'))
            lamdanf = float('%.3f' % (lnf / Df))
            self.main.ed_lamdanf.setText(str(lamdanf))

        except ValueError:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Неправильный ввод данных.")
            msg.setWindowTitle("Ошибка")
            msg.setDetailedText("Вводить можно только цифры. Числа должны быть положительными.")
            msg.exec_()

        if permision_for_curve[0] and permision_for_curve[1] and permision_for_curve[2]:
            self.main.tabData.setTabEnabled(1, True)


    # РАСЧЕТ ДАННЫХ ГОНДОЛ ДВИГАТЕЛЯ И ШАССИ
    def CalculateGondola(self):
        global c_lamda_el_for_cxa

        try:
            # Удлинение гондола двигателя
            global lgd, Dgd, lamdagd
            lgd = float(self.main.ed_lgd.text().replace(',', '.'))
            Dgd = float(self.main.ed_Dgd.text().replace(',', '.'))
            lamdagd = float('%.3f' % (lgd / Dgd))
            self.main.ed_lamdagd.setText(str(lamdagd))

            # Удлинение смоченная поверхность гондола двигателя
            global Ssm_gd
            Ssm_gd = float('%.3f' % (2.5 * lgd * Dgd))
            self.main.ed_Ssmgd.setText(str(Ssm_gd))

            # Удлинение удлинение носовой части гондола двигателя
            global ln_gd, lamdan_gd
            ln_gd = float(self.main.ed_lngd.text().replace(',', '.'))
            lamdan_gd = float('%.3f' % (ln_gd / Dgd))
            self.main.ed_lamdangd.setText(str(lamdan_gd))

            # Удлинение гондола шасси
            global lgsh, Dgsh, lamdagsh
            lgsh = float(self.main.ed_lgsh.text().replace(',', '.'))
            Dgsh = float(self.main.ed_Dgsh.text().replace(',', '.'))
            lamdagsh = float('%.3f' % (lgsh / Dgsh))
            self.main.ed_lamdagsh.setText(str(lamdagsh))

            # Удлинение смоченная поверхность гондола шасси
            global Ssm_gsh
            Ssm_gsh = float('%.3f' % (2.5 * lgsh * Dgsh))
            self.main.ed_Ssmgsh.setText(str(Ssm_gsh))

            # Удлинение удлинение носовой части гондола шасси
            global ln_gsh, lamdan_gsh
            ln_gsh = float(self.main.ed_lngsh.text().replace(',', '.'))
            lamdan_gsh = float('%.3f' % (ln_gsh / Dgsh))
            self.main.ed_lamdangsh.setText(str(lamdan_gsh))

            global ngd, ngsh
            ngd = float(self.main.ed_ngd.text().replace(',', '.'))
            ngsh = float(self.main.ed_dsh.text().replace(',', '.'))

        except ValueError:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Неправильный ввод данных.")
            msg.setWindowTitle("Ошибка")
            msg.setDetailedText("Вводить можно только цифры. Числа должны быть положительными.")
            msg.exec_()
        except ZeroDivisionError:
            lamdagd = lamdan_gd = lamdagsh = lamdan_gsh = Ssm_gd = Ssm_gsh = 0
            self.main.ed_lamdagd.setText(str(lamdagd))
            self.main.ed_lamdangd.setText(str(lamdan_gd))
            self.main.ed_Ssmgsh.setText(str(Ssm_gsh))
            self.main.ed_Ssmgd.setText(str(Ssm_gd))
            self.main.ed_lamdagsh.setText(str(lamdagsh))
            self.main.ed_lamdangsh.setText(str(lamdan_gsh))

            c_lamda_el_for_cxa.append(lamdagd)
            c_lamda_el_for_cxa.append(lamdagsh)

        if permision_for_curve[0] and permision_for_curve[1] and permision_for_curve[2]:
            self.main.tabData.setTabEnabled(1, True)


    # СОХРАНЕНИЕ ОБЩИХ ДАННЫХ
    def SafeCommonData(self):

        try:
            global Gvzl, V, number, P0, H, Dv, Fv
            Dv = float(self.main.ed_Dv.text().replace(',', '.'))
            Fv = float('%.3f' % (math.pi * Dv * Dv / 4))
            Gvzl = float(self.main.ed_Gvzl.text().replace(',', '.'))
            V = float(self.main.ed_V.text())
            number = float(self.main.ed_number.text().replace(',', '.'))
            P0 = float(self.main.ed_P0.text().replace(',', '.'))
            H = float(self.main.ed_H.text().replace(',', '.'))
            self.find_with_H(H)

            # Тип двигателей
            global type
            if self.main.radioButton_PD.isChecked():
                type = 'ТВД и ПД'
            elif self.main.radioButton_TRD.isChecked():
                type = 'ТРД'

            if type != '':
                global Gm
                Gm = self.findGm()

            global Kint, cxk_light
            if self.main.rb_circle.isChecked():
                Kint = self.request_Kint('Низкоплан с фюзеляжем круглого сечения')
            elif self.main.rb_oval.isChecked():
                Kint = self.request_Kint('Низкоплан с фюзеляжем овального сечения')
            elif self.main.rb_rectangle.isChecked():
                Kint = self.request_Kint('Низкоплан с фюзеляжем прямоугольного сечения')
            elif self.main.rb_middle.isChecked():
                Kint = self.request_Kint('Среднеплан')
            elif self.main.rb_high.isChecked():
                Kint = self.request_Kint('Высокоплан')

            if self.main.rb_1.isChecked():
                cxk_light = self.request_cxk('Фонарь с плоской передней стенкой')
            elif self.main.rb_2.isChecked():
                cxk_light = self.request_cxk('Плавный переход задней части фонаря в фюзеляж')
            elif self.main.rb_3.isChecked():
                cxk_light = self.request_cxk('Фонарь обтекаемой формы с плоскими передними стенками')
            elif self.main.rb_no.isChecked():
                cxk_light = self.request_cxk('Отсутствует')

            global nlight
            nlight = float(self.main.ed_nlight.text())

            permision_for_curve[2] = True
        except ValueError:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Неправильный ввод данных.")
            msg.setWindowTitle("Ошибка")
            msg.setDetailedText("Вводить можно только цифры. Числа должны быть положительными.")
            msg.exec_()

        if permision_for_curve[0] and permision_for_curve[1] and permision_for_curve[2]:
            self.main.tabData.setTabEnabled(1, True)
            self.MakeMkr()


    # РАСЧЕТ И ПОСТРОЕНИЕ КРИВОЙ ЗАВИСИМОСТИ Мкр
    def MakeMkr(self):
        global permission
        permission[0]='+'
        permissiom_for_polar[0]='+'
        self.iconbutton(self.main.buMkr, self.button_curve)
        self.main.buCre.setEnabled(True)
        for p in permission:
            if p=='':
                self.main.buUp.setEnabled(False)
                self.main.buUp.setIcon(QtGui.QIcon('images/x.svg'))
                self.main.buDown.setEnabled(False)
                self.main.buDown.setIcon(QtGui.QIcon('images/x.svg'))
        self.main.buHelp.setEnabled(True)
        self.main.tabWidget.setCurrentIndex(0)

        # ЗАПОЛНЕНИЕ МАССИВОВ ДАННЫМИ
        global linear_size, c_lamda_el_for_cxa, Sk_el_for_cxa, n_el_for_cxa
        linear_size = [b, bgo, bvo, bp, lf, lgd, lgsh, 0]  # фонарь!=0
        c_lamda_el_for_cxa = [c_, cgo_, cvo_, cp_, lamdaf, lamdagd, lamdagsh, 0]
        Sk_el_for_cxa = [S, Sgo, Svo, Sp, Ssm * 0.5, Ssm_gd * 0.5, Ssm_gsh * 0.5, Smf]
        n_el_for_cxa = [1, ngo, nvo, npylon, 1, ngd, ngsh, nlight]

        # Расчет координат вспомогательной прямой
        arr_cya = [0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7]
        arr_Mkr = []
        for cya in arr_cya:
            expression = 1 - 0.445 * math.pow(lamdaef - 1, 1 / 9) * (0.175 + 3.25 * c_) * (
                    math.cos(x_degree * math.pi / 180) + (0.365 * cya ** 2) / math.cos(
                x_degree * math.pi / 180) ** 5)
            arr_Mkr.append(float('%.3f' % expression))

        global list_Mkr
        list_Mkr = arr_Mkr.copy()

        # Занесение координат в БД
        arr = []
        arr_ = [0, 1]
        for i in range(8):
            arr_[0] = arr_cya[i]
            arr_[1] = arr_Mkr[i]
            arr.append(tuple(arr_))
        cor = tuple(arr)
        self.cursor.execute('DELETE FROM "Кривая зависимость Мкр";', )
        sqlite_insert_query = """INSERT INTO 'Кривая зависимость Мкр'
                                         (Cya, Мкр) VALUES (?, ?);"""
        self.cursor.executemany(sqlite_insert_query, cor)
        self.connection.commit()

        # Отображение в приложении
        SQLquery = 'SELECT * FROM "Кривая зависимость Мкр"'

        tablecol = 1
        self.main.tableWidget_Mkr.setRowCount(2)
        self.main.tableWidget_Mkr.setColumnCount(9)
        for row in self.cursor.execute(SQLquery):
            for col in range(2):
                item = QtWidgets.QTableWidgetItem(str(row[col]))
                tablerow = col
                item.setTextAlignment(QtCore.Qt.AlignCenter)
                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                self.main.tableWidget_Mkr.setItem(tablerow, tablecol, item)
            tablecol = tablecol + 1

        item1 = QtWidgets.QTableWidgetItem('Cya')
        item2 = QtWidgets.QTableWidgetItem('Мкр')
        self.main.tableWidget_Mkr.setItem(0, 0, item1)
        self.main.tableWidget_Mkr.setItem(1, 0, item2)

        # Расчет точки ?????
        global Gpol, Mrasch, cya_rasch
        Vms = float('%.3f' % (V / 3.6))  # км/ч в м/с
        Mrasch = float('%.4f' % (Vms / aH))  # определение расчетного числа маха
        Gpol = Gvzl - 0.5 * Gm
        cya_rasch = float('%.4f' % ((2 * Gpol * g) / (pH * Vms * Vms * S)))

        # Отображение найденных переменных
        self.main.la_Mrasch_Mkr.setText(str(Mrasch))
        self.main.la_Gpol_Mkr.setText(str(Gpol))
        self.main.la_Cyarasch_Mkr.setText(str(cya_rasch))

        # Отрисовка
        self.figure.clear()
        ax = self.figure.add_subplot(111)

        tck, u = interpolate.splprep([arr_cya, arr_Mkr], s=0)
        xnew, ynew = interpolate.splev(np.linspace(0, 1, 100), tck, der=0)
        ax.plot(arr_cya, arr_Mkr, '.', xnew, ynew, color='k') # отрисовка

        ax.plot(cya_rasch, Mrasch, marker='.', color='navy')
        ax.text(cya_rasch, Mrasch, '  A(Суа расч;Мрасч)', rotation=0, fontsize=7)

        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.set_xlabel('$\it{Cya}$', loc='right', fontsize=11)
        ax.set_ylabel('$\it{Mкр}$', loc='top', rotation=0)

        self.figure.tight_layout()
        self.canvas.draw()

        self.figure.savefig('graphics/Mkr.png')

        try:
            self.ExportData()
        except PermissionError:
            self.ErrorMessage()


    # ПОСТРОЕНИЕ ВСПОМОГАТЕЛЬНОЙ КРИВОЙ cya = f(a)
    def MakeHelp(self):
        global permission
        permission[1]="+"
        self.main.buUp.setEnabled(True)
        self.main.buDown.setEnabled(True)
        self.main.buUp.setIcon(QtGui.QIcon('images/caret-right.svg'))
        self.main.buDown.setIcon(QtGui.QIcon('images/caret-right.svg'))
        self.iconbutton(self.main.buHelp, self.button_curve)
        self.main.tabWidget.setCurrentIndex(1)
        self.fhelp.clear()
        ax = self.fhelp.add_subplot(111)

        # Линейный участок, характеризующий безотрывное обтекания крыла
        # Точка 1 - (а0, 0)

        # Точка 2 - (а, суа)
        alfa_px2 = 5  # произвольно
        cya_py2 = float('%.3f' % (caya * (alfa_px2 - a0)))

        # Точка 4 - (а, суа_max)
        # Коэф cya max
        global Kn, Cyamax
        Kn = float('%.3f' % self.find_Kn(n))
        self.find_Cyamaxprof()
        Cyamax = float('%.3f' % (Cyamaxprof * Kn * (1 + math.cos(x_degree * math.pi / 180)) / 2))
        k = cya_py2 / (alfa_px2 - a0)
        b = (-1) * k * a0
        px4 = float('%.3f' % ((Cyamax - b) / k))

        print('Kn = ' + str(Kn))
        print('Cyamax = ' + str(Cyamax))
        # Точка 3 - (x, 0.85 * суа_max)
        cya_py3 = float('%.3f' % (0.85 * Cyamax))
        px3 = float('%.3f' % ((cya_py3 - b) / k))

        # Точка 5 - (deltaalfa ,суа_max)
        deltaalfa = 2  # !!! произвольно
        px5 = px4 + deltaalfa
        py5 = Cyamax

        # заполнение списков для последующего использования во вспомогательной поляре
        global a_list_help, cya_list_help
        a_list_help.clear()
        cya_list_help.clear()
        a_list_help.append(a0)
        cya_list_help.append(0)
        a = 1  # !!!!!
        while px5 - a > 3:
            a_list_help.append(a)
            cya_list_help.append(float('%.3f' % (k*a+b)))
            a += 3
        a_list_help.append(px5)
        cya_list_help.append(Cyamax)

        # Отрисовка
        arrX_ = [a0, alfa_px2, px3, px4]
        arrY_ = [0, cya_py2, cya_py3, Cyamax]
        ax.plot(arrX_, arrY_, '--', color='k')
        arrX = [a0, alfa_px2, px3]
        arrY = [0, cya_py2, cya_py3]
        ax.plot(arrX, arrY, color='k')

        px = (px3 + px4) / 2
        py = (cya_py3 + Cyamax) / 2
        # Точка 6 - вспомогательная для дуги
        px6 = 2 * px5 - px3
        py6 = cya_py3

        arr_intr_X = [px3, px, px5, px6]
        arr_intr_Y = [cya_py3, py, py5, py6]
        tck, u = interpolate.splprep([arr_intr_X, arr_intr_Y], s=0)
        xnew, ynew = interpolate.splev(np.linspace(0, 1, 100), tck, der=0)
        ax.plot(arr_intr_X, arr_intr_Y, ' ', xnew, ynew, color='k')

        ax.plot(a0, 0, marker='o')
        ax.plot(alfa_px2, cya_py2, marker='o')
        ax.plot(px3, cya_py3, marker='o')
        ax.plot(px4, Cyamax, marker='o')
        ax.plot(px5, py5, marker='o')

        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.set_ylabel('$\it{Cya}$', loc = 'top', rotation=0)
        ax.set_xlabel('$\it{α, °}$', loc = 'right', fontsize=11)

        self.fhelp.tight_layout()
        self.chelp.draw()
        self.fhelp.savefig('graphics/Help.png')

        # Отображение в окне
        self.main.la_1_help.setText(str(a0) + '; ' + str(0))
        self.main.la_2_help.setText(str(alfa_px2) + '; ' + str(cya_py2))
        self.main.la_3_help.setText(str(px3) + '; ' + str(cya_py3))
        self.main.la_4_help.setText(str(px4) + '; ' + str(Cyamax))
        self.main.la_5_help.setText(str(px5) + '; ' + str(Cyamax))

        self.main.la_Re_help.setText(str(Re))
        self.main.la_Vmin_help.setText(str(float('%.3f' % Vmin)))
        self.main.la_Cyamaxprof_help.setText(str(Cyamaxprof))
        self.main.la_Kn_help.setText(str(Kn))
        self.main.la_cya_max.setText(str(Cyamax))

        try:
            wb = openpyxl.load_workbook('Расчет самолета.xlsx')
            # заполнение второго листа с данными
            st = wb['Кривые']
            st.cell(27,2).value = Cyamax
            st.cell(28,2).value = Cyamaxprof
            st.cell(29,2).value = Re
            st.cell(30,2).value = Kn
            st.cell(31,2).value = Vmin

            st.cell(27,8).value = a0
            st.cell(28,8).value = alfa_px2
            st.cell(29,8).value = px3
            st.cell(30,8).value = px4
            st.cell(31,8).value = px5
            st.cell(27, 9).value = 0
            st.cell(28, 9).value = cya_py2
            st.cell(29, 9).value = cya_py3
            st.cell(30, 9).value = Cyamax
            st.cell(31, 9).value = Cyamax

            wb.save('Расчет самолета.xlsx')
        except PermissionError:
            self.ErrorMessage()

        # x=np.array( [px3,px5])
        # y = np.array([cya_py3, Cyamax])
        # x_new = np.linspace(x.min(), x.max(), 500)
        # f = interpolate.interp1d(x, y, kind='quadratic')
        # y_smooth = f(x_new)
        # ax.plot(x_new, y_smooth)

    # ПОИСК КОЭФФИЦИЕНТА ОТ ТИПА ЗАКРЫЛОК
    def find_zak(self, Flap_type):
        global dCyamax, dCxomax
        Sql_request = 'SELECT * FROM Закрылки WHERE Тип_механизации = %s' % '"' + Flap_type + '"'
        self.cursor.execute(Sql_request)
        cur = self.cursor.fetchone()
        dCyamax = float(cur[2])
        dCxomax = float(cur[3])


    def cal_system_x(self, x1, y1, x2, y2, y):
        k = (y1 - y2) / (x1 - x2)
        b_koef = y1 - k * x1
        return float('%.5f' % ((y-b_koef)/k))

    def cal_k_b_koef(self, x1, y1, x2, y2):
        k_koef = (y1 - y2) / (x1 - x2)
        b_koef = y1 - k_koef * x1
        return [k_koef, b_koef]

    # ПОСТРОЕНИЕ ВЗЛЕТНОЙ КРИВОЙ (С УЧЕТОМ И БЕЗ УЧЕТА ВЛИЯНИЯ ЭКРАНА ЗЕМЛИ)
    def MakeUp(self):
        self.main.tabWidget.setCurrentIndex(2)
        self.iconbutton(self.main.buUp, self.button_curve)

        # 1) Без учета влияния экрана земли
        # Определение переменных по графикам da0_vzl
        self.find_zak(Flap_type)
        da0_vzl = self.call_da0(deltavzl)  # определение приращения угла атаки нулевой подъемной силы в радианах
        # Преращение коэф подъем силы от выпуска предкрылков
        dCyamax_pr = 0.6 * Sobpr_

        # Преращение коэф подъем силы от выпуска закрылков
        dCyamax_zak_vzl = float('%.3f' % (
                4.83 * dCyamax * Sobzak_ * abs(da0_vzl) * math.cos(xshzak * math.pi / 180) * math.cos(
            xshzak * math.pi / 180)))

        # УРАВНЕНИЯ ПРИ НАЛИЧИИ ВИНТОВ
        if Dv != 0:
            W = Vmin/2 + math.sqrt(math.pow(Vmin/2, 2) + (4*Gvzl)/(math.pi*n*p_zero*10*Dv*Dv))
            dCya_obd = -caya*(-1.5+a0)*math.pow(W/Vmin, 2)*Sobd_
            da0_obd = (1.5+a0)*math.pow(W/Vmin, 2)*Sobd_
        else:
            dCya_obd = 0
            da0_obd = 0

        # Максимальный коэф подъемной силы при взлете
        global Cyamax_vzl
        Cyamax_vzl = float('%.3f' % (Cyamax + dCyamax_pr + dCyamax_zak_vzl+dCya_obd)) #+dCya_obd
        a0_vzl = float('%.3f' % (a0 + (da0_vzl+da0_obd) * 180 / math.pi))  # +da0_obd

        # Точки
        a1 = a0_vzl
        cya1 = 0

        a2 = 5
        cya2 = float('%.3f' % (caya * (5 - a0_vzl)))

        cya3 = 0.85*Cyamax_vzl
        a3 = self.cal_system_x(a1, cya1, a2, cya2, cya3)

        a4 = self.cal_system_x(a1, cya1, a2, cya2, Cyamax_vzl)
        cya4 = Cyamax_vzl

        amax = float('%.3f' % (a4+2))

        # Отрисовка
        self.fUp.clear()
        ax = self.fUp.add_subplot(111)
        # Прямая
        points_a = [a1, a2, a3]
        points_Cya = [cya1, cya2, cya3]
        # Вспомогательные точки для дуги
        px = (a3 + a4) / 2
        py = (cya3 + cya4) / 2
        px2 = 2 * amax - a3
        py2 = cya3
        arr_intr_X = [a3, px, amax, px2]
        arr_intr_Y = [cya3, py, Cyamax_vzl, py2]

        tck, u = interpolate.splprep([arr_intr_X, arr_intr_Y], s=0)
        xnew, ynew = interpolate.splev(np.linspace(0, 1, 100), tck)
        ax.plot(arr_intr_X, arr_intr_Y, ' ', xnew, ynew, color='k')

        ax.plot([a1, a3], [cya1, cya3], marker='', color='k')
        ax.plot(amax, Cyamax_vzl, marker='o', color='tab:green')

        # Отображение найденных переменных
        self.main.la_Cyamaxvzl_Up.setText(str(Cyamax_vzl))
        self.main.la_a0vzl_Up.setText(str(a0_vzl))
        self.main.lineEdit_3.setText(str('('+ str(a1) + '; ' + str(cya1) + ')'))
        self.main.lineEdit_2.setText(str('('+ str(amax)+ '; ' + str(Cyamax_vzl) + ')'))

        # заполнение списков для взлетной кривой
        global a_list_up, cya_list_up
        a_list_up.clear()
        cya_list_up.clear()
        a_list_up.append(a0_vzl)
        cya_list_up.append(0)
        angle = int((abs(a0_vzl)+abs(amax))/6)
        koef = self.cal_k_b_koef(a1, cya1, a2, cya2)
        a = a0_vzl
        for i in range(6):
            a = float('%.0f' % (angle+a))
            if Cyamax_vzl < koef[0] * a + koef[1]:
                a = a-2
            a_list_up.append(a)
            cya_list_up.append(float('%.3f' % (koef[0] * a + koef[1])))
        a_list_up.append(amax)
        cya_list_up.append(Cyamax_vzl)

        # 2) С учетом влияния экрана земли
        # Приращение КПС, вызванный экранным влиянияем земли
        dCyamax_zak_vzl_scrin = float('%.3f' % (-1 * 0.115 * math.exp(-0.5 * hvzl / bsrzak) * Cyamax_vzl))

        # КПС, вызванный экранным влиянияем земли
        global Cyamax_vzl_scrin, lamdaef_scrin
        Cyamax_vzl_scrin = float('%.3f' % (Cyamax_vzl + dCyamax_zak_vzl_scrin))

        # Фиктивное удлинение крыла, учитывающее влияние экрана земли
        lamdaef_scrin = float('%.3f' % ((lamdaef / 2.23) * ((math.pi * l) / (8 * hvzl) + 2)))

        # Производная с учетом влияния экрана земли
        expression = (2 * math.pi * lamdaef_scrin * math.cos(x_degree * math.pi / 180)) / (
                57.3 * (lamdaef_scrin + 2 * math.cos(x_degree * math.pi / 180)))
        caya_scrin = float('%.3f' % expression)

        # Точки
        cya2_s = float('%.3f' % (expression * (5 - a0_vzl)))

        cya3_s = 0.85 * Cyamax_vzl_scrin
        a3_s = self.cal_system_x(a1, cya1, a2, cya2_s, cya3_s)

        a4_s = self.cal_system_x(a1, cya1, a2, cya2_s, Cyamax_vzl_scrin)
        cya4_s = Cyamax_vzl_scrin

        amax_s = float('%.3f' % (a4_s + 2))

        # Прямая
        points_a_s = [a1,a2, a3_s]
        points_Cya_s = [cya1, cya2_s, cya3_s]
        # Вспомогательные точки для дуги
        px = (a3_s + a4_s) / 2
        py = (cya3_s + cya4_s) / 2
        px2 = 2 * amax_s - a3_s
        py2 = cya3_s
        arr_intr_X_s = [a3_s, px, amax_s, px2]
        arr_intr_Y_s = [cya3_s, py, Cyamax_vzl_scrin, py2]

        # Отрисовка
        tck, u = interpolate.splprep([arr_intr_X_s, arr_intr_Y_s], s=0)
        xnew_s, ynew_s = interpolate.splev(np.linspace(0, 1, 100), tck)
        ax.plot(arr_intr_X_s, arr_intr_Y_s, ' ', xnew_s, ynew_s, color='tab:blue')

        ax.plot([a1, a3_s], [cya1, cya3_s], marker='', color='tab:blue')
        ax.plot(amax_s, Cyamax_vzl_scrin, marker='o', color='tab:red')
        ax.plot(a1, cya1, marker='o', color='gold')

        #ax.set_title('Взлетные кривые Cya = f(a)', loc='right', pad=5, fontsize=11)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.set_ylabel('$\it{Cya}$', loc='top', rotation=0)
        ax.set_xlabel('$\it{α, °}$', loc='right', fontsize=11)
        self.fUp.tight_layout()
        self.cUp.draw()
        self.fUp.savefig('graphics/Up.png')

        # Отображение найденных переменных
        self.main.la_Cyamaxvzl_scrin_Up.setText(str(Cyamax_vzl_scrin))
        self.main.la_Caya_scrin_Up.setText(str(caya_scrin))
        self.main.la_lamda_scrin_Up.setText(str(lamdaef_scrin))
        self.main.lineEdit.setText(str('(' + str(amax_s) + '; ' + str(Cyamax_vzl_scrin) + ')'))


        # заполнение списков для взлетной кривой
        global a_list_up_scrin, cya_list_up_scrin
        a_list_up_scrin.clear()
        cya_list_up_scrin.clear()
        a_list_up_scrin.append(a0_vzl)
        cya_list_up_scrin.append(0)
        angle = int((abs(a0_vzl) + abs(amax_s)) / 6)
        koef = self.cal_k_b_koef(a1, cya1, a2, cya2_s)
        a = a0_vzl
        for i in range(6):
            a = float('%.0f' % (angle + a))
            if Cyamax_vzl_scrin < koef[0] * a + koef[1]:
                a = a-2
            a_list_up_scrin.append(a)
            cya_list_up_scrin.append(float('%.3f' % (koef[0] * a + koef[1])))
        a_list_up_scrin.append(amax_s)
        cya_list_up_scrin.append(Cyamax_vzl_scrin)

        try:
            wb = openpyxl.load_workbook('Расчет самолета.xlsx')
            # заполнение второго листа с данными
            st = wb['Кривые']
            # без экрана
            st.cell(52, 2).value = dCyamax
            st.cell(53, 2).value = da0_vzl
            st.cell(55, 3).value = dCyamax_pr
            st.cell(56, 3).value = dCyamax_zak_vzl
            st.cell(57, 3).value = Cyamax_vzl
            st.cell(58, 3).value = a0_vzl

            l_a = points_a.copy()
            l_a.append(amax)
            l_cya = points_Cya.copy()
            l_cya.append(Cyamax_vzl)
            for col in range(8, 10):
                i = 0
                for row in range(55, 59):
                    cell = st.cell(row, col)
                    if col == 8:
                        cell.value = l_a[i]
                    else:
                        cell.value = l_cya[i]
                    i += 1

            # с экраном
            st.cell(62, 3).value = hvzl / bsrzak
            st.cell(63, 3).value = caya_scrin
            st.cell(64, 3).value = lamdaef_scrin
            st.cell(66, 3).value = dCyamax_zak_vzl_scrin
            st.cell(67, 3).value = Cyamax_vzl_scrin
            st.cell(68, 3).value = a0_vzl

            l_as = points_a_s.copy()
            l_as.append(amax_s)
            l_cyas = points_Cya_s.copy()
            l_cyas.append(Cyamax_vzl_scrin)
            for col in range(8, 10):
                i = 0
                for row in range(66, 70):
                    cell = st.cell(row, col)
                    if col == 8:
                        cell.value = l_as[i]
                    else:
                        cell.value = l_cyas[i]
                    i += 1

            wb.save('Расчет самолета.xlsx')
        except:
            ErrorMessage()

    # ПОСТРОЕНИЕ ПОСАДОЧНЫХ КРИВЫХ (С УЧЕТОМ И БЕЗ УЧЕТА ВЛИЯНИЯ ЭКРАНА ЗЕМЛИ)
    def MakeDown(self):
        self.main.tabWidget.setCurrentIndex(3)
        self.iconbutton(self.main.buDown, self.button_curve)

        # 1) Без учета влияния экрана земли
        #    Определение переменных
        da0_pos = self.call_da0(deltapos)
        dCyamax_pr = 0.6 * Sobpr_  # преращение коэф подъем силы от выпуска предкрылков

        # Приращение КПС при выпущенных закрылках при посадке
        dCya_max_zak_pos = float(
            '%.3f' % (4.83 * dCyamax * Sobzak_ * abs(da0_pos) * (math.cos(xshzak * math.pi / 180) ** 2)))

        # КПС при посадке
        global Cya_max_pos
        Cya_max_pos = float('%.3f' % (Cyamax + dCyamax_pr + dCya_max_zak_pos))
        a0_pos = float('%.3f' % (a0 + da0_pos * 180 / math.pi))

        # Точка
        # Точки
        a1 = a0_pos
        cya1 = 0

        a2 = 5
        cya2 = float('%.3f' % (caya * (5 - a0_pos)))

        cya3 = 0.85 * Cya_max_pos
        a3 = self.cal_system_x(a1, cya1, a2, cya2, cya3)

        a4 = self.cal_system_x(a1, cya1, a2, cya2, Cya_max_pos)
        cya4 = Cya_max_pos

        amax = float('%.3f' % (a4 + 2))

        # Прямая
        points_a = [a1, a2, a3]
        points_Cya = [cya1, cya2, cya3]
        # Вспомогательные точки для дуги
        px = (a3 + a4) / 2
        py = (cya3 + cya4) / 2
        px2 = 2 * amax - a3
        py2 = cya3
        arr_intr_X = [a3, px, amax, px2]
        arr_intr_Y = [cya3, py, cya4, py2]

        # Отрисовка
        self.fDown.clear()  # отчистка графика
        ax = self.fDown.add_subplot(111)

        tck, u = interpolate.splprep([arr_intr_X, arr_intr_Y], k=2)
        xnew, ynew = interpolate.splev(np.linspace(0, 1, 100), tck, der=0)
        ax.plot(arr_intr_X, arr_intr_Y, ' ', xnew, ynew, color='k')

        ax.plot([a1, a3], [cya1, cya3], marker=' ', color='k')
        ax.plot(amax, Cya_max_pos, marker='o', color='tab:green')

        #ax.set_title('Посадочные кривые Cya = f(a)', loc='right', pad=5, fontsize=11)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.set_ylabel('$\it{Cya}$', loc='top', rotation=0)
        ax.set_xlabel('$\it{α, °}$', loc='right', fontsize=11)
        self.fDown.tight_layout()

        # Отображение найденных переменных
        self.main.la_Cyamaxpos_Down.setText(str(Cya_max_pos))
        self.main.la_a0pos_Down.setText(str(a0_pos))
        self.main.lineEdit_5.setText(str('(' + str(a1) + '; ' + str(cya1) + ')'))
        self.main.lineEdit_4.setText(str('(' + str(amax) + '; ' + str(Cya_max_pos) + ')'))

        # заполнение списков для взлетной кривой
        global a_list_down, cya_list_down
        a_list_down.clear()
        cya_list_down.clear()
        a_list_down.append(a0_pos)
        cya_list_down.append(0)
        angle = int((abs(a0_pos) + abs(amax)) / 6)
        koef = self.cal_k_b_koef(a1, cya1, a2, cya2)
        a = a0_pos
        for i in range(6):
            a = float('%.0f' % (angle + a))
            if Cya_max_pos < koef[0] * a + koef[1]:
                a = a-2
            a_list_down.append(a)
            cya_list_down.append(float('%.3f' % (koef[0] * a + koef[1])))
        a_list_down.append(amax)
        cya_list_down.append(Cya_max_pos)

        # 2) C учетом влияния экрана земли
        #    определение переменных
        x = x_degree * math.pi / 180
        h_pos = hpos / bsrzak

        # определение производной caya
        lamda_scrin = float('%.3f' % ((lamdaef / 2.23) * ((math.pi * l) / (8 * hpos) + 2)))
        caya_scrin = float(
            '%.3f' % ((2 * math.pi * lamda_scrin * math.cos(x)) / (57.3 * (lamda_scrin + 2 * math.cos(x)))))

        # макс КПС силы при посадке с учетом экрана земли с выпущенными закрылками
        dCya_max_zak_pos_scrin = float('%.3f' % (-0.115 * math.exp(-0.5 * h_pos) * Cya_max_pos))

        # макс КПС силы при посадке с учетом экрана земли
        global Cya_max_pos_scrin, lamdaef_down_scrin
        Cya_max_pos_scrin = float('%.3f' % (Cya_max_pos + dCya_max_zak_pos_scrin))
        lamdaef_down_scrin=lamda_scrin

        # Точки
        cya2_s = float('%.3f' % (caya_scrin * (5 - a0_pos)))

        cya3_s = 0.85 * Cya_max_pos_scrin
        a3_s = self.cal_system_x(a1, cya1, a2, cya2_s, cya3_s)

        a4_s = self.cal_system_x(a1, cya1, a2, cya2_s, Cya_max_pos_scrin)
        cya4_s = Cya_max_pos_scrin

        amax_s = float('%.3f' % (a4_s + 2))

        # Прямая
        points_a_s = [a1, a2, a3_s]
        points_Cya_s = [cya1, cya2_s, cya3_s]
        # Вспомогательные точки для дуги
        px = (a3_s + a4_s) / 2
        py = (cya3_s + cya4_s) / 2
        px2 = 2 * amax_s - a3_s
        py2 = cya3_s
        arr_intr_X_s = [a3_s, px, amax_s, px2]
        arr_intr_Y_s = [cya3_s, py, Cya_max_pos_scrin, py2]

        tck, u = interpolate.splprep([arr_intr_X_s, arr_intr_Y_s], s=0)
        xnew, ynew = interpolate.splev(np.linspace(0, 1, 100), tck)
        ax.plot(arr_intr_X_s, arr_intr_Y_s, ' ', xnew, ynew, color='tab:blue')

        ax.plot([a1, a3_s], [cya1, cya3_s], marker=' ', color='tab:blue')
        ax.plot(amax_s, Cya_max_pos_scrin, marker='o', color='tab:red')
        ax.plot(a1, cya1, marker='o', color='gold')
        self.cDown.draw()
        self.fDown.savefig('graphics/Down.png')
        # Отображение найденных переменных
        self.main.la_Cyamaxpos_scrin_Down.setText(str(Cya_max_pos_scrin))
        self.main.la_Caya_scrin_Down.setText(str(caya_scrin))
        self.main.la_lamda_scrin_Down.setText(str(lamda_scrin))
        self.main.lineEdit_6.setText(str('(' + str(amax_s) + '; ' + str(Cya_max_pos_scrin) + ')'))

        # заполнение списков для взлетной кривой
        global a_list_down_scrin, cya_list_down_scrin
        a_list_down_scrin.clear()
        cya_list_down_scrin.clear()
        a_list_down_scrin.append(a0_pos)
        cya_list_down_scrin.append(0)
        angle = int((abs(a0_pos) + abs(amax_s)) / 6)
        koef = self.cal_k_b_koef(a1, cya1, a2, cya2_s)
        a = a0_pos
        for i in range(6):
            a = float('%.0f' % (angle + a))
            if Cya_max_pos_scrin < koef[0] * a + koef[1]:
                a = a-2
            a_list_down_scrin.append(a)
            cya_list_down_scrin.append(float('%.3f' % (koef[0] * a + koef[1])))
        a_list_down_scrin.append(amax_s)
        cya_list_down_scrin.append(Cya_max_pos_scrin)

        try:
            wb = openpyxl.load_workbook('Расчет самолета.xlsx')
            # заполнение второго листа с данными
            st = wb['Кривые']
            # без экрана
            st.cell(102, 2).value = da0_pos
            st.cell(106, 3).value = dCya_max_zak_pos
            st.cell(107, 3).value = Cya_max_pos
            st.cell(108, 3).value = a0_pos

            l_a = points_a.copy()
            l_a.append(amax)
            l_cya = points_Cya.copy()
            l_cya.append(Cya_max_pos)
            for col in range(8, 10):
                i = 0
                for row in range(105, 109):
                    cell = st.cell(row, col)
                    if col == 8:
                        cell.value = l_a[i]
                    else:
                        cell.value = l_cya[i]
                    i += 1

            # с экраном
            st.cell(112, 3).value = h_pos
            st.cell(113, 3).value = caya_scrin
            st.cell(114, 3).value = lamdaef_down_scrin
            st.cell(116, 3).value = dCya_max_zak_pos_scrin
            st.cell(117, 3).value = Cya_max_pos_scrin
            st.cell(118, 3).value = a0_pos

            l_as = points_a_s.copy()
            l_as.append(amax_s)
            l_cyas = points_Cya_s.copy()
            l_cyas.append(Cya_max_pos_scrin)
            for col in range(8, 10):
                i = 0
                for row in range(116, 120):
                    cell = st.cell(row, col)
                    if col == 8:
                        cell.value = l_as[i]
                    else:
                        cell.value = l_cyas[i]
                    i += 1

            wb.save('Расчет самолета.xlsx')
        except:
            self.ErrorMessage()

    # Выбор чисел Маха от типа двигателя
    def func_type(self, type):
        if type == 'ТРД':
            list_M = [float('%.3f' % Mrasch), 0, 0.7, 0.8, 0.85, 0.9, 0.95]
            return list_M
        elif type == 'ТВД и ПД':
            list_M = [float('%.3f' % Mrasch), 0, 0.4, 0.5, 0.6, 0.7]
            return list_M

    # ПОСТРОЕНИЕ КРЕЙСЕРСКИХ КРИВЫХ
    def MakeCruise(self):
        self.iconbutton(self.main.buCre, self.button_curve)
        for p in permission:
            if p=='':
                self.main.buUp.setEnabled(False)
                self.main.buUp.setIcon(QtGui.QIcon('images/x.svg'))
                self.main.buDown.setEnabled(False)
                self.main.buDown.setIcon(QtGui.QIcon('images/x.svg'))
        self.main.tabWidget.setCurrentIndex(4)

        list_M = self.func_type(type)
        list_caya_szh = []
        list_cya = []
        for M in list_M:
            caya_szh = caya / math.sqrt(1 - M * M)
            list_caya_szh.append(float('%.3f' % caya_szh))

            cya = caya_szh * (5 - a0)
            list_cya.append(float('%.3f' % cya))

        #  Построение
        self.fCruise.clear()
        ax = self.fCruise.add_subplot(111)
        for cya in list_cya:
            point_alfa = [a0, 5]
            point_cya = [0, cya]
            ax.plot(point_alfa, point_cya, color='k', linewidth=1)
            ax.annotate('M = '+ str(list_M[list_cya.index(cya)]), xy=(5, cya), xytext = (5,5),  textcoords='offset points')


        ax.vlines(5, 0, list_cya[-1], color='k', linewidth=1, linestyle='--')
        #ax.set_title('Крейсерские кривые Суа = f(a)', loc='left', pad=5, fontsize=11)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.set_ylabel('$\it{Cya}$', loc='top', rotation=0)
        ax.set_xlabel('$\it{α, °}$', loc='right', fontsize=11)
        self.fCruise.tight_layout()
        # Занесение координат в БД
        arr = []
        arr_ = [0, 1, 2]
        for i in range(len(list_M)):
            arr_[0] = list_M[i]
            arr_[1] = list_caya_szh[i]
            arr_[2] = list_cya[i]
            arr.append(tuple(arr_))
        cor = tuple(arr)
        self.cursor.execute('DELETE FROM "Крейсерская кривая";', )
        sqlite_insert_query = """INSERT INTO 'Крейсерская кривая'
                                                (M, Caya_szh, Cya) VALUES (?, ?, ?);"""
        self.cursor.executemany(sqlite_insert_query, cor)
        self.connection.commit()

        # Отображение в таблице
        SQLquery = 'SELECT * FROM "Крейсерская кривая"'

        tablecol = 1
        self.main.tableWidget_Cruise.setRowCount(3)
        self.main.tableWidget_Cruise.setColumnCount(len(list_M) + 1)
        for row in self.cursor.execute(SQLquery):
            for col in range(3):
                item = QtWidgets.QTableWidgetItem(str(row[col]))
                tablerow = col
                item.setTextAlignment(QtCore.Qt.AlignCenter)
                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                self.main.tableWidget_Cruise.setItem(tablerow, tablecol, item)
            tablecol = tablecol + 1

        item1 = QtWidgets.QTableWidgetItem('М')
        item2 = QtWidgets.QTableWidgetItem('Суа сж')
        item3 = QtWidgets.QTableWidgetItem('Суа')
        self.main.tableWidget_Cruise.setItem(0, 0, item1)
        self.main.tableWidget_Cruise.setItem(1, 0, item2)
        self.main.tableWidget_Cruise.setItem(2, 0, item3)

        self.main.tabData.setTabEnabled(2, True)
        #self.MakeHelpPolyr()
        self.cCruise.draw()
        self.fCruise.savefig('graphics/Cre.png')
        try:

            wb = openpyxl.load_workbook('Расчет самолета.xlsx')
            # заполнение второго листа с данными
            st = wb['Кривые']
            st.cell(151, 2).value = Mrasch
            if type =='ТРД':
                for row in range(155, 157):
                    i=0
                    for col in range(2, 9):
                        cell = st.cell(row, col)
                        if row ==155:
                            cell.value = list_caya_szh[i]
                        else:
                            cell.value = list_cya[i]
                        i+=1
            else:
                st.cell(154,8).value = None
                for col in range(4,8):
                    st.cell(154, col).value = list_M[col-2]
                for row in range(155, 157):
                    i=0
                    for col in range(2, 8):
                        cell = st.cell(row, col)
                        if row ==155:
                            cell.value = list_caya_szh[i]
                        else:
                            cell.value = list_cya[i]
                        i+=1

            wb.save('Расчет самолета.xlsx')
        except: self.ErrorMessage()
        self.MakeHelpPolyr()

    # ВСПОМОГАТЕЛЬНАЯ ПОЛЯРА
    def MakeHelpPolyr(self):
        permision_for_export[0] = '1'
        self.main.tabPolyr.setCurrentIndex(0)
        self.iconbutton(self.main.buHelpPolyr, self.button_polar)
        for p in permision_for_export:
            if p == '':
                self.main.buExport.setIcon(QtGui.QIcon('images/x.svg'))

        print('ВСПОМОГАТЕЛЬНАЯ ПОЛЯРА')

        # РАСЧЕТ КОЭФФИЦИЕНТА ПРОФИЛЬНОГО СОПРОТИВЛЕНИЯ
        Vmin_pol = float('%.3f' % (math.sqrt((2 * Gpol * g) / (p_zero * S * Cyamax))))
        # print('Vmin_pol = ' + str(Vmin_pol))
        M = Vmin_pol/340.294 # ЧЕ ЗА КОНСТАНТА ???????????
        delta = self.call_delta(lamdaef, n)
        print('delta = '+ str(delta))

        print('Кинт = ' + str(Kint))
        print('cxk = ' + str(cxk_light))

        global xtau_el_for_cxa
        Re_el_for_cxa = []
        xtau_el_for_cxa = []
        cf_el_for_cxa = []
        nc_el_for_cxa = []
        nm_el_for_cxa = []
        nint_el_for_cxa = []
        cxk_el_for_cxa = []
        chislitel_el_for_cxa = []

        for el in linear_size:
            index = linear_size.index(el)
            # число Рейнольдса
            Re_el = float('%.3f' % ((float(Vmin_pol) * float(el)) / (v_zero * math.pow(10, 6))))  # степень 10^6
            Re_el_for_cxa.append(Re_el)
            # точка перехода
            if index == 0:
                xtau_el_for_cxa.append(xtau_)
                nint_el_for_cxa.append(float('%.3f' % (1 - Kint * Sf_)))  # nинт для крыла
            else:
                xtau_el_for_cxa.append(0)
            if el != 0:
                # коэф сопротивления плоской пластине
                cf_el_for_cxa.append(self.find_2cf(xtau_el_for_cxa[index], Re_el))
                # коэф nМ
                # коэф nc
                if index < 4:
                    nc_el_for_cxa.append(self.find_nc_wing(xtau_el_for_cxa[index], c_lamda_el_for_cxa[index]))
                    nm_el_for_cxa.append(self.call_nM_wing(c_lamda_el_for_cxa[index],M))
                else:
                    nc_el_for_cxa.append(self.find_nc_body_rotate(c_lamda_el_for_cxa[index]))
                    nm_el_for_cxa.append(self.call_nM_body_rotation(c_lamda_el_for_cxa[index], M))
                # коэф nинт
                if index != 0:
                    nint_el_for_cxa.append(1)
                # коэф cxk
                cxk = cf_el_for_cxa[index] * nc_el_for_cxa[index] * nm_el_for_cxa[index] * nint_el_for_cxa[index]
                cxk_el_for_cxa.append(float('%.5f' % cxk))
            else:
                nc_el_for_cxa.append(0)
                cf_el_for_cxa.append(0)
                nm_el_for_cxa.append(0)
                nint_el_for_cxa.append(0)
                cxk_el_for_cxa.append(0)

        cxk_el_for_cxa[-1]=cxk_light
        global S_proiz, cxo
        for el in Sk_el_for_cxa:
            i = Sk_el_for_cxa.index(el)
            chislitel = n_el_for_cxa[i] * cxk_el_for_cxa[i] * el
            chislitel_el_for_cxa.append(float('%.5f' % chislitel))
            S_proiz.append(float('%.5f' % (n_el_for_cxa[i] * nc_el_for_cxa[i]*nint_el_for_cxa[i]*el)))

        cxo = float('%.5f' % (sum(chislitel_el_for_cxa) * 1.04 / S))

        th = ['Крыло', 'Горизонтальное оперение', 'Вертикальное оперение', 'Пилон', 'Фюзеляж',
              'Гондола двигателя', 'Гондола шасси', 'Фонарь, кабины пилотов']
        td = ['Линейный размер', 'Re', 'xt', '2cf', 'c_, lamda', 'nc',
              'nM', 'nинт', 'cxk', 'Sk', 'n', 'n*cxk*ck*Sk']

        table = PrettyTable(th)

        table.add_row(linear_size)
        table.add_row(Re_el_for_cxa)
        table.add_row(xtau_el_for_cxa)
        table.add_row(cf_el_for_cxa)
        table.add_row(c_lamda_el_for_cxa)
        table.add_row(nc_el_for_cxa)
        table.add_row(nm_el_for_cxa)
        table.add_row(nint_el_for_cxa)
        table.add_row(cxk_el_for_cxa)
        table.add_row(Sk_el_for_cxa)
        table.add_row(n_el_for_cxa)
        table.add_row(chislitel_el_for_cxa)
        table.add_column('Расчетная величина', td)

        print(table)  # Печатаем таблицу
        print('cxo = ' + str(cxo))

        global nc_for_cre, nint_for_cre
        nc_for_cre = nc_el_for_cxa.copy()
        nint_for_cre = nint_el_for_cxa.copy()

        # Приращение коэффициента профильного сопротивления
        cya__list = []
        dCxp_list = []
        # Коэффициент вихревого индуктивного сопротивления самолета
        Cxi_list = []
        # Коэффициент лобового сопротивления
        Cxa_list=[]
        self.fP_Help.clear()
        ax = self.fP_Help.add_subplot()
        kmax=0

        for el in cya_list_help:
            # Отношение КПС к максимальному КПС
            cya_ = el/Cyamax
            cya__list.append(float('%.4f' % (cya_)))
            # Приращение коэффициента профильного сопротивления
            exp_dCxp = math.pow(cya_, 4)*(1-math.exp(-0.1*math.pow((cya_ - 0.4), 2)))
            dCxp_list.append(float('%.4f' % exp_dCxp))
            # Коэффициент вихревого индуктивного сопротивления
            exp_Cxi = ((el*el)/(math.pi*lamdaef))*((1+delta)/math.sqrt(1-M*M))
            Cxi_list.append(float('%.4f' % exp_Cxi))
            # Коэффициент лобового сопротивления
            Cxa_list.append(float('%.4f' % (cxo+exp_dCxp+exp_Cxi)))
        # Отрисовка с помощью интерполяции
        tck, u = interpolate.splprep([Cxa_list, cya_list_help], s=0)
        xnew, ynew = interpolate.splev(np.linspace(0, 1, 100), tck)
        ax.plot(Cxa_list, cya_list_help, '.', xnew, ynew, color='tab:blue')
        for x in xnew:  # построение касательной к поляре
            k = ynew[np.where(xnew == x)]/x
            if kmax < k:
                kmax = k
                xmax=x
        ax.plot(xmax, kmax*xmax, '.', [0, xmax*2], [0, kmax*xmax*2 ], linewidth = 1, color = 'crimson')


        # КОЛИЧЕТСВО ДВИГАТЕЛЕЙ ЭТО НУМЕР !!!!!!!!!

        table_coordinats = PrettyTable(a_list_help)
        table_coordinats.add_row(cya_list_help)
        table_coordinats.add_row(cya__list)
        table_coordinats.add_row(dCxp_list)
        table_coordinats.add_row(Cxi_list)
        table_coordinats.add_row(Cxa_list)

        print(table_coordinats)
        self.main.la_Vminpol.setText(str(Vmin_pol))
        self.main.la_cxo.setText(str(cxo))
        self.main.la_M_Vmin.setText(str(float('%.3f' % M)))

        index=0
        for i, j in zip(Cxa_list, cya_list_help):
            ax.annotate(str(a_list_help[index]), xy=(i, j))
            index +=1

        # point_x = [min(Cxa_list),min(Cxa_list)]
        # point_y = [0, max(cya_list_help)/2]
        # ax.plot(point_x, point_y, color='green')
        # ax.set_title('Вспомогательная поляра', loc='left', pad=5, fontsize=11)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.set_ylabel('$\it{Cya}$', loc='top', rotation=0)
        ax.set_xlabel('$\it{Cxa}$', loc='right', fontsize=11)
        self.fP_Help.tight_layout()
        self.cP_Help.draw()
        self.fP_Help.savefig('graphics/P_help.png')

        try:
            wb = openpyxl.load_workbook('Расчет самолета.xlsx')
            # заполнение второго листа с данными
            if type == 'ТРД':
                st = wb['Поляры']
            else:
                st = wb['Поляры ТВД']
            st.cell(18, 2).value = cxo
            st.cell(20, 2).value = Vmin_pol
            st.cell(21, 2).value = M
            st.cell(20, 6).value = delta

            for col in range(2, 10):
                st.cell(4, col).value = linear_size[col - 2]
                st.cell(5, col).value = Re_el_for_cxa[col - 2]
                st.cell(6, col).value = xtau_el_for_cxa[col - 2]
                st.cell(7, col).value = cf_el_for_cxa[col - 2]
                st.cell(8, col).value = c_lamda_el_for_cxa[col - 2]
                st.cell(9, col).value = nc_el_for_cxa[col - 2]
                st.cell(10, col).value = nm_el_for_cxa[col - 2]
                st.cell(11, col).value = nint_el_for_cxa[col - 2]
                st.cell(12, col).value = cxk_el_for_cxa[col - 2]
                st.cell(13, col).value = Sk_el_for_cxa[col - 2]
                st.cell(14, col).value = n_el_for_cxa[col - 2]
                st.cell(15, col).value = chislitel_el_for_cxa[col - 2]
            for col in range(2, len(a_list_help) + 2):
                st.cell(24, col).value = a_list_help[col - 2]
                st.cell(25, col).value = cya_list_help[col - 2]
                st.cell(26, col).value = cya__list[col - 2]
                st.cell(27, col).value = dCxp_list[col - 2]
                st.cell(28, col).value = Cxi_list[col - 2]
                st.cell(29, col).value = Cxa_list[col - 2]

            wb.save('Расчет самолета.xlsx')
        except:
            self.ErrorMessage()

    # ВЗЛЕТНАЯ ПОЛЯРА
    def MakeUpPolyr(self):
        permision_for_export[1] = '1'
        print()
        print("ВЗЛЕТНАЯ ПОЛЯРА ")
        print(dCxomax)
        print(self.find_dCxo_zak(bzak_, deltavzl))
        print(deltavzl)
        self.main.tabPolyr.setCurrentIndex(1)
        self.iconbutton(self.main.buUpPolyr, self.button_polar)
        for p in permision_for_export:
            if p == '':
                self.main.buExport.setIcon(QtGui.QIcon('images/x.svg'))
                self.main.buExport.setEnabled(False)
            else:
                self.main.buExport.setEnabled(True)

        # 1) без учета влияния экрана земли
        delta = self.call_delta(lamdaef, n)
        cya__list_up = []
        dCxp_list_up = []
        Cxi_list_up = []
        Cxa_list_up = []


        cxo_vzl =  cxo + 0.5*cxo + 10 * self.find_dCxo_zak(bzak_, deltavzl) * Sobzak_ * dCxomax
        print('cxo_vzl = '+ str(float('%.5f' % cxo_vzl)))
        Vvzl = math.sqrt((2*Gvzl*g)/(0.8 * p_zero * S * Cyamax_vzl_scrin)) #!!!!! почему с экраном ?
        print('V_vzl = '+ str(float('%.3f' % Vvzl)))
        Mvzl = Vvzl/340.294 # CONST
        print('Mvzl = '+ str(float('%.3f' % Mvzl)))

        for el in cya_list_up:
            #
            cya_ = el /Cyamax_vzl # какой суа макс нужен ??
            cya__list_up.append(float('%.4f' % (cya_)))
            #
            exp_dCxp = math.pow(cya_, 4) * (1 - math.exp(-0.1 * math.pow((cya_ - 0.4), 2)))
            dCxp_list_up.append(float('%.4f' % exp_dCxp))
            #
            exp_Cxi = ((el * el) / (math.pi * lamdaef)) * ((1 + delta) / math.sqrt(1 - Mvzl * Mvzl))
            Cxi_list_up.append(float('%.4f' % exp_Cxi))
            #
            Cxa_list_up.append(float('%.4f' % (cxo_vzl + exp_dCxp + exp_Cxi)))

        table_coordinats = PrettyTable(a_list_up)
        table_coordinats.add_row(cya_list_up)
        table_coordinats.add_row(cya__list_up)
        table_coordinats.add_row(dCxp_list_up)
        table_coordinats.add_row(Cxi_list_up)
        table_coordinats.add_row(Cxa_list_up)
        print(table_coordinats)

        # 2) c учетом влияния экрана земли
        delta = self.call_delta(lamdaef_scrin, n)
        cya__list_up_scrin = []
        dCxp_list_up_scrin = []
        Cxi_list_up_scrin = []
        Cxa_list_up_scrin = []

        Vvzl_scrin = math.sqrt((2 * Gvzl * g) / (0.8 * p_zero * S * Cyamax_vzl_scrin))  # !!!!! почему с экраном ?
        print('V_vzl_scrin = ' + str(float('%.3f' % Vvzl_scrin)))
        Mvzl_scrin = Vvzl_scrin / 340.294  # CONST
        print('Mvzl_scrin = ' + str(float('%.3f' % Mvzl_scrin)))

        for el in cya_list_up_scrin:
            #
            cya_ = el / Cyamax_vzl_scrin  # какой суа макс нужен ??
            cya__list_up_scrin.append(float('%.4f' % (cya_)))
            #
            exp_dCxp = math.pow(cya_, 4) * (1 - math.exp(-0.1 * math.pow((cya_ - 0.4), 2)))
            dCxp_list_up_scrin.append(float('%.4f' % exp_dCxp))
            #
            exp_Cxi = ((el * el) / (math.pi * lamdaef_scrin)) * ((1 + delta) / math.sqrt(1 - Mvzl_scrin * Mvzl_scrin))
            Cxi_list_up_scrin.append(float('%.4f' % exp_Cxi))
            #
            Cxa_list_up_scrin.append(float('%.4f' % (cxo_vzl + exp_dCxp + exp_Cxi)))

        table_coordinats_scrin = PrettyTable(a_list_up_scrin)
        table_coordinats_scrin.add_row(cya_list_up_scrin)
        table_coordinats_scrin.add_row(cya__list_up_scrin)
        table_coordinats_scrin.add_row(dCxp_list_up_scrin)
        table_coordinats_scrin.add_row(Cxi_list_up_scrin)
        table_coordinats_scrin.add_row(Cxa_list_up_scrin)
        print(table_coordinats_scrin)

        # отрисовка
        self.fP_Up.clear()  # отчистка графика
        ax = self.fP_Up.add_subplot()
        # без экрана
        tck, u = interpolate.splprep([Cxa_list_up, cya_list_up], s=0)
        xnew, ynew = interpolate.splev(np.linspace(0, 1, 100), tck)
        ax.plot(Cxa_list_up, cya_list_up, '.', xnew, ynew, color='k')

        kmax_up = xmax = ymax = -1
        for x in xnew:  # построение касательной к поляре
            k = ynew[np.where(xnew == x)]/x
            if kmax_up < k:
                kmax_up = k
                xmax = x
                ymax = ynew[np.where(xnew == xmax)]
        self.main.la_k_vzl.setText(str(float('%.3f' % kmax_up)) + ' при (' + str(float('%.3f' % xmax)) + '; ' + str(float('%.3f' % ymax))+ ')')

        index=0
        for i, j in zip(Cxa_list_up, cya_list_up):
            ax.annotate(str(a_list_up[index]), xy=(i, j))
            index +=1

        # с экраном
        tck, u = interpolate.splprep([Cxa_list_up_scrin, cya_list_up_scrin], s=0)
        xnew, ynew = interpolate.splev(np.linspace(0, 1, 100), tck)
        ax.plot(Cxa_list_up_scrin, cya_list_up_scrin, '.', xnew, ynew, color='tab:blue')

        kmax_up = xmax = ymax = -1
        for x in xnew:  # построение касательной к поляре
            k = ynew[np.where(xnew == x)] / x
            if kmax_up < k:
                kmax_up = k
                xmax = x
                ymax = ynew[np.where(xnew == xmax)]
        self.main.la_k_vzl_s.setText(str(float('%.3f' % kmax_up)) + ' при (' + str(float('%.3f' % xmax)) + '; ' + str(float('%.3f' % ymax)) + ')')

        index = 0
        for i, j in zip(Cxa_list_up_scrin, cya_list_up_scrin):
            ax.annotate(str(a_list_up_scrin[index]), xy=(i, j))
            index += 1

        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.set_ylabel('$\it{Cya}$', loc='top', rotation=0)
        ax.set_xlabel('$\it{Cxa}$', loc='right', fontsize=11)
        self.fP_Up.tight_layout()
        self.cP_Up.draw()
        self.fP_Up.savefig('graphics/P_Up.png')

        self.main.la_Vvzl.setText(str(float('%.3f' % Vvzl)))
        self.main.la_M_vzl.setText(str(float('%.3f' % Mvzl)))
        self.main.la_cxo_vzl.setText(str(float('%.3f' % cxo_vzl)))

        try:

            wb = openpyxl.load_workbook('Расчет самолета.xlsx')
            # заполнение второго листа с данными
            if type == 'ТРД':
                st = wb['Поляры']
            else:
                st = wb['Поляры ТВД']
            st.cell(48, 3).value = dCxomax
            st.cell(49, 3).value = float('%.4f' % self.find_dCxo_zak(bzak_, deltavzl))
            st.cell(50, 2).value = Mvzl
            st.cell(51, 2).value = Vvzl
            st.cell(48, 7).value = float('%.4f' % (0.5*cxo))
            st.cell(49, 7).value = float('%.4f' % (10 * self.find_dCxo_zak(bzak_, deltavzl) * Sobzak_ * dCxomax))
            st.cell(50, 7).value = cxo_vzl

            for col in range(2, len(a_list_up) + 2):
                st.cell(54, col).value = a_list_up[col - 2]
                st.cell(55, col).value = cya_list_up[col - 2]
                st.cell(56, col).value = cya__list_up[col - 2]
                st.cell(57, col).value = dCxp_list_up[col - 2]
                st.cell(58, col).value = Cxi_list_up[col - 2]
                st.cell(59, col).value = Cxa_list_up[col - 2]

            for col in range(2, len(a_list_up_scrin) + 2):
                st.cell(64, col).value = a_list_up_scrin[col - 2]
                st.cell(65, col).value = cya_list_up_scrin[col - 2]
                st.cell(66, col).value = cya__list_up_scrin[col - 2]
                st.cell(67, col).value = dCxp_list_up_scrin[col - 2]
                st.cell(68, col).value = Cxi_list_up_scrin[col - 2]
                st.cell(69, col).value = Cxa_list_up_scrin[col - 2]

            wb.save('Расчет самолета.xlsx')
        except:
            self.ErrorMessage()

    # ПОСАДОЧНАЯ ПОЛЯРА
    def MakeDownPolyr(self):
        # print(dCxomax)
        # print(self.find_dCxo_zak(bzak_, deltavzl))
        # print(deltavzl)
        permision_for_export[2] = '1'
        self.main.tabPolyr.setCurrentIndex(2)
        self.iconbutton(self.main.buDownPolyr, self.button_polar)
        for p in permision_for_export:
            if p == '':
                self.main.buExport.setIcon(QtGui.QIcon('images/x.svg'))
                self.main.buExport.setEnabled(False)
            else:
                self.main.buExport.setEnabled(True)

        print()
        print("ПОСАДОЧНАЯ ПОЛЯРА ")
        print(self.find_dCxo_zak(bzak_, deltapos))

        # 1) без учета влияния экрана земли
        delta = self.call_delta(lamdaef, n)
        cya__list_down = []
        dCxp_list_down = []
        Cxi_list_down = []
        Cxa_list_down = []


        cxo_down =  cxo + 0.5*cxo + 10 * self.find_dCxo_zak(bzak_, deltapos) * Sobzak_ * dCxomax #!!!!!!!!!!!!!!!
        print('cxo = '+str(cxo))
        print('Dcxo_zak_pos = '+str(10 * self.find_dCxo_zak(bzak_, deltapos) * Sobzak_ * dCxomax))

        print('cxo_down = '+ str(float('%.3f' % cxo_down)))
        Vdown = math.sqrt((2*Gpol*g)/(0.8 * p_zero * S * Cya_max_pos_scrin)) #!!!!! почему с экраном ?
        print('V_down = '+ str(float('%.3f' % Vdown)))
        Mdown = Vdown/340.294 # CONST
        print('Mdown = '+ str(float('%.3f' % Mdown)))

        for el in cya_list_down:
            #
            cya_ = el /Cya_max_pos # какой суа макс нужен ??
            cya__list_down.append(float('%.4f' % (cya_)))
            #
            exp_dCxp = math.pow(cya_, 4) * (1 - math.exp(-0.1 * math.pow((cya_ - 0.4), 2)))
            dCxp_list_down.append(float('%.4f' % exp_dCxp))
            #
            exp_Cxi = ((el * el) / (math.pi * lamdaef)) * ((1 + delta) / math.sqrt(1 - Mdown * Mdown))
            Cxi_list_down.append(float('%.4f' % exp_Cxi))
            #
            Cxa_list_down.append(float('%.4f' % (cxo_down + exp_dCxp + exp_Cxi)))

        table_coordinats = PrettyTable(a_list_down)
        table_coordinats.add_row(cya_list_down)
        table_coordinats.add_row(cya__list_down)
        table_coordinats.add_row(dCxp_list_down)
        table_coordinats.add_row(Cxi_list_down)
        table_coordinats.add_row(Cxa_list_down)
        print(table_coordinats)

        # 1) c учетом влияния экрана земли
        delta = self.call_delta(lamdaef_down_scrin, n)
        cya__list_down_scrin = []
        dCxp_list_down_scrin = []
        Cxi_list_down_scrin = []
        Cxa_list_down_scrin = []

        Vdown_scrin = math.sqrt((2 * Gpol * g) / (0.8 * p_zero * S * Cya_max_pos_scrin))  # !!!!! почему с экраном ?
        print('V_down_scrin = ' + str(float('%.3f' % Vdown_scrin)))
        Mdown_scrin = Vdown_scrin / 340.294  # CONST
        print('Mdown_scrin = ' + str(float('%.3f' % Mdown_scrin)))

        for el in cya_list_down_scrin:
            #
            cya_ = el / Cya_max_pos_scrin  # какой суа макс нужен ??
            cya__list_down_scrin.append(float('%.4f' % (cya_)))
            #
            exp_dCxp = math.pow(cya_, 4) * (1 - math.exp(-0.1 * math.pow((cya_ - 0.4), 2)))
            dCxp_list_down_scrin.append(float('%.4f' % exp_dCxp))
            #
            exp_Cxi = ((el * el) / (math.pi * lamdaef_down_scrin)) * ((1 + delta) / math.sqrt(1 - Mdown_scrin * Mdown_scrin))
            Cxi_list_down_scrin.append(float('%.4f' % exp_Cxi))
            #
            Cxa_list_down_scrin.append(float('%.4f' % (cxo_down + exp_dCxp + exp_Cxi)))

        table_coordinats_scrin = PrettyTable(a_list_down_scrin)
        table_coordinats_scrin.add_row(cya_list_down_scrin)
        table_coordinats_scrin.add_row(cya__list_down_scrin)
        table_coordinats_scrin.add_row(dCxp_list_down_scrin)
        table_coordinats_scrin.add_row(Cxi_list_down_scrin)
        table_coordinats_scrin.add_row(Cxa_list_down_scrin)
        print(table_coordinats_scrin)

        # отрисовка
        self.fP_Down.clear()  # отчистка графика
        ax = self.fP_Down.add_subplot()
        # без экрана
        tck, u = interpolate.splprep([Cxa_list_down, cya_list_down], k=2, s = 0)
        xnew, ynew = interpolate.splev(np.linspace(0, 1, 100), tck)
        ax.plot(Cxa_list_down, cya_list_down, '.', xnew, ynew, color='k')

        kmax_down = xmax = ymax = -1
        for x in xnew:  # построение касательной к поляре
            k = ynew[np.where(xnew == x)] / x
            if kmax_down < k:
                kmax_down = k
                xmax = x
                ymax = ynew[np.where(xnew == xmax)]
        self.main.la_k_pos.setText(str(float('%.3f' % kmax_down)) + ' при (' + str(float('%.3f' % xmax)) + '; ' + str(float('%.3f' % ymax)) + ')')

        index=0
        for i, j in zip(Cxa_list_down, cya_list_down):
            ax.annotate(str(a_list_down[index]), xy=(i, j))
            index +=1

        # с экраном
        tck, u = interpolate.splprep([Cxa_list_down_scrin, cya_list_down_scrin], k =2, s=0)
        xnew, ynew = interpolate.splev(np.linspace(0, 1, 100), tck)
        ax.plot(Cxa_list_down_scrin, cya_list_down_scrin, '.', xnew, ynew, color='tab:blue')

        kmax_down = xmax = ymax = -1
        for x in xnew:  # построение касательной к поляре
            k = ynew[np.where(xnew == x)] / x
            if kmax_down < k:
                kmax_down = k
                xmax = x
                ymax = ynew[np.where(xnew == xmax)]
        self.main.la_k_pos_s.setText(str(float('%.3f' % kmax_down)) + ' при (' + str(float('%.3f' % xmax)) + '; ' + str(float('%.3f' % ymax)) + ')')

        index = 0
        for i, j in zip(Cxa_list_down_scrin, cya_list_down_scrin):
            ax.annotate(str(a_list_down_scrin[index]), xy=(i, j))
            index += 1

        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.set_ylabel('$\it{Cya}$', loc='top', rotation=0)
        ax.set_xlabel('$\it{Cxa}$', loc='right', fontsize=11)
        self.fP_Down.tight_layout()
        self.cP_Down.draw()
        self.fP_Down.savefig('graphics/P_Down.png')

        self.main.la_Vpos.setText(str(float('%.3f' % Vdown)))
        self.main.la_M_pos.setText(str(float('%.3f' % Mdown)))
        self.main.la_cxo_pos.setText(str(float('%.3f' % cxo_down)))
        try:
            wb = openpyxl.load_workbook('Расчет самолета.xlsx')
            # заполнение второго листа с данными
            if type == 'ТРД':
                st = wb['Поляры']
            else:
                st = wb['Поляры ТВД']
            st.cell(94, 3).value = float('%.4f' % self.find_dCxo_zak(bzak_, deltapos))
            st.cell(95, 2).value = Mdown
            st.cell(96, 2).value = Vdown
            st.cell(94, 7).value = cxo
            st.cell(95, 7).value = float('%.4f' % (10 * self.find_dCxo_zak(bzak_, deltapos) * Sobzak_ * dCxomax))
            st.cell(96, 7).value = cxo_down


            for col in range(2, len(a_list_down) + 2):
                st.cell(100, col).value = a_list_down[col - 2]
                st.cell(101, col).value = cya_list_down[col - 2]
                st.cell(102, col).value = cya__list_down[col - 2]
                st.cell(103, col).value = dCxp_list_down[col - 2]
                st.cell(104, col).value = Cxi_list_down[col - 2]
                st.cell(105, col).value = Cxa_list_down[col - 2]

            for col in range(2, len(a_list_down_scrin) + 2):
                st.cell(110, col).value = a_list_down_scrin[col - 2]
                st.cell(111, col).value = cya_list_down_scrin[col - 2]
                st.cell(112, col).value = cya__list_down_scrin[col - 2]
                st.cell(113, col).value = dCxp_list_down_scrin[col - 2]
                st.cell(114, col).value = Cxi_list_down_scrin[col - 2]
                st.cell(115, col).value = Cxa_list_down_scrin[col - 2]
            wb.save('Расчет самолета.xlsx')
        except:
            self.ErrorMessage()

    # КРЕЙСЕРСКИЕ ПОЛЯРЫ
    def MakeCruisePolyr(self):
        permision_for_export[3] = '1'
        self.main.tabPolyr.setCurrentIndex(3)
        self.iconbutton(self.main.buCrePolyr, self.button_polar)
        for p in permision_for_export:
            if p == '':
                self.main.buExport.setIcon(QtGui.QIcon('images/x.svg'))
                self.main.buExport.setEnabled(False)
            else:
                self.main.buExport.setEnabled(True)

        self.fP_Cre.clear()  # отчистка графика
        ax = self.fP_Cre.add_subplot()
        global list_K
        list_M = self.func_type(type)
        list_M.pop(0)
        # print(list_M)

        list_Re = []
        list_Xak=[]
        list_2cf=[]
        list_nM =[]
        list_cxo_cruise = []
        list_lamdan = c_lamda_el_for_cxa.copy()
        list_lamdan[4]=lamdanf
        list_lamdan[5]=lamdan_gd
        list_lamdan[6]=lamdan_gsh

        #TRYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYY
        wb = openpyxl.load_workbook('Расчет самолета.xlsx')
        if type == 'ТРД':
            st = wb['Поляры']
        else:
            st = wb['Поляры ТВД']
        arr = [linear_size, list_lamdan, xtau_el_for_cxa, nc_for_cre, nint_for_cre, Sk_el_for_cxa, n_el_for_cxa,
               S_proiz]
        i1 = 0
        for row in range(141, 149):
            for col in range(2, 10):
                list_row = arr[i1]
                st.cell(row, col).value = list_row[col - 2]
            i1 += 1

        row = 148
        #
        for M in list_M:
            list_Re.clear()
            list_Xak.clear()
            list_2cf.clear()
            list_nM.clear()
            for el in linear_size:
                if el != 0:
                    index = linear_size.index(el)
                    el_c_lamda = list_lamdan[index]
                    xt = xtau_el_for_cxa[index]
                    if M == 0:
                        Re = (Vmin * el) / (vH * math.pow(10, 6))
                    else:
                        Re = (M * aH * el) / (vH * math.pow(10, 6))
                    list_Re.append(float('%.3f' % Re))
                    list_2cf.append(self.find_2cf(xt, Re))
                    if index < 4:
                        list_nM.append(self.call_nM_wing(el_c_lamda * 100, M))
                    else:
                        list_nM.append(self.call_nM_body_rotation(el_c_lamda, M))
                    Xak = list_2cf[index] * S_proiz[index] * list_nM[index]
                    list_Xak.append(float('%.3f' % Xak))
                else:
                    list_Xak.append(0)
                    list_Re.append(0)
                    list_2cf.append(0)
                    list_nM.append(0)
            # t = PrettyTable(['Крыло', "Гор оперение", "Вер оперение", "Пилон", "Фюзеляж", "ГД", "ГШ", "Фонарь"])
            # t.add_row(list_Re)
            # t.add_row(list_2cf)
            # t.add_row(list_nM)
            # t.add_row(list_Xak)
            # print()
            # print('M = ' + str(M))
            # print(t)
            #
            row += 2
            if row == 150:
                st.cell(row, 5).value = Vmin
            else:
                st.cell(row, 5).value = float('%.4f' % (M * aH))
            row += 2
            arr.clear()
            arr = [list_Re, list_2cf, list_nM, list_Xak]
            i2=0
            for r in range(row, row + 4):
                for col in range(2, 10):
                    list_row = arr[i2]
                    st.cell(r, col).value = list_row[col - 2]
                i2 += 1
            #
            cxo_cruise = float('%.5f' % (sum(list_Xak) * 1.04 / S))
            list_cxo_cruise.append(cxo_cruise)
            row += 4
            st.cell(row, 2).value = cxo_cruise
            # print('cxo = ' + str(cxo_cruise))
            cos = math.cos(xc_degree)
            # Максимальный коэф волнового сопротивления при числе Маха =
        Mc_xvo_max = math.pow(cos, -1) * (1 + 0.4 * (math.pow(c_, 3 / 2) / math.pow(cos, 2 / 3)) *
                                          (2 - lamdaef * math.pow(c_ * cos * cos, 1 / 3)))
        cxvo_max = (2 * math.pi * lamdaef * c_ * c_ * cos) / (2 + lamdaef * math.pow(c_, 1 / 3) * math.pow(cos, 5 / 3))
        # print('M max = ' + str(float('%.3f' % Mc_xvo_max)))
        # print('C max = ' + str(float('%.3f' % cxvo_max)))
        if type=='ТРД':
            st.cell(200, 3).value = Mc_xvo_max
            st.cell(201, 3).value = cxvo_max
        else:
            st.cell(192, 3).value = Mc_xvo_max
            st.cell(193, 3).value = cxvo_max

        # расчет координат поляр
        list_cya = [0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7]
        list_cxi = []
        list_cxa = []
        list_cxvo = []
        list_cxvi = []
        list_Am=[]
        delta = self.call_delta(lamdaef, n)
        # print(delta)
        l_M0 = []
        l_M7 = []
        l_M8 = []
        l_M85 = []
        l_M9 = []
        l_M95 = []
        l_M4 = []
        l_M5 = []
        l_M6 = []
        list_cxa_POL = []
        if type == 'ТРД':
            row = 201
        else:
            row = 192
        for cya in list_cya:
            Mkr = list_Mkr[list_cya.index(cya)]
            list_cxi.clear()
            list_cxa.clear()
            list_cxvo.clear()
            list_cxvi.clear()
            list_Am.clear()
            q=0 #########
            for M in list_M:
                i = list_M.index(M)
                # коэффициент вихревого сопротивления
                cxi = math.pow(cya, 2) / (math.pi * lamdaef) * (1 + delta) / math.sqrt(1 - M * M)
                list_cxi.append(float('%.5f' % cxi))
                if type == 'ТРД':
                    if M > Mkr:
                        # составляющая коэф волнового сопротивления, не зависящая от cya
                        Am = (M - Mkr) / (Mc_xvo_max - Mkr)
                        list_Am.append(float('%.5f' % Am))
                        cxvo = cxvo_max * math.pow(Am, 3) * (4 - 3 * Am)
                        list_cxvo.append(float('%.5f' % cxvo))
                        # коэф волнового сопротивления
                        cxvi = 25 * lamdaef * math.pow(c_, 1 / 3) * math.pow(M - Mkr, 3) * cxi
                        list_cxvi.append(float('%.5f' % cxvi))
                    else:
                        list_cxvo.append(0)
                        list_cxvi.append(0)
                        list_Am.append(0)
                    # КЛС
                    cxa = list_cxo_cruise[i] + list_cxi[i] + list_cxvo[i] + list_cxvi[i]
                    list_cxa.append(float('%.5f' % cxa))
                    # СДЕЛАНО ПО ТУПОМУ!!!!!!!!!!!!!!
                    if M == 0:
                        l_M0.append(cxa)
                    elif M == 0.7:
                        l_M7.append(cxa)
                    elif M == 0.8:
                        l_M8.append(cxa)
                    elif M == 0.85:
                        l_M85.append(cxa)
                    elif M == 0.9:
                        l_M9.append(cxa)
                    elif M == 0.95:
                        l_M95.append(cxa)
                else:
                    # КЛС
                    cxa = list_cxo_cruise[i] + list_cxi[i]
                    list_cxa.append(float('%.5f' % cxa))
                    if M == 0:
                        l_M0.append(cxa)
                    elif M == 0.4:
                        l_M4.append(cxa)
                    elif M == 0.5:
                        l_M5.append(cxa)
                    elif M == 0.6:
                        l_M6.append(cxa)
                    elif M == 0.7:
                        l_M7.append(cxa)
            if type == 'ТРД':
                list_cxa_POL = [l_M0, l_M7, l_M8, l_M85, l_M9, l_M95]
                row += 3
                st.cell(row, 2).value = Mkr
                arr.clear()
                arr = [list_cxo_cruise, list_cxi, list_Am, list_cxvo, list_cxvi, list_cxa]
                for col in range(4, 10):
                    p = 0
                    for r in range(row, row + 6):
                        list_col = arr[q]
                        st.cell(r, col).value = list_col[p]
                        p += 1
                    q += 1
                row += 5

                # t = PrettyTable()
                # t.add_column("M", list_M)
                # t.add_column("cxo", list_cxo_cruise)
                # t.add_column("cxi", list_cxi)
                # t.add_column('cxvo', list_cxvo)
                # t.add_column('cxvi', list_cxvi)
                # t.add_column("cxa", list_cxa)
                # print()
                # print('cya = ' + str(cya))
                # print('Mkr = ' + str(Mkr))
                # print(t)
            else:
                list_cxa_POL = [l_M0, l_M4, l_M5, l_M6, l_M7]
                self.main.buFindK.setEnabled(False)
                self.main.cb_H.setEnabled(False)
                self.main.cb_M.setEnabled(False)
                row += 4
                st.cell(row, 2).value = Mkr
                arr.clear()
                arr = [list_cxo_cruise, list_cxi, list_cxa]
                for col in range(4, 7):
                    p = 0
                    for r in range(row, row + 5):
                        list_col = arr[q]
                        st.cell(r, col).value = list_col[p]
                        p += 1
                    q += 1
                row += 4

                # t = PrettyTable()
                # t.add_column("M", list_M)
                # t.add_column("cxo", list_cxo_cruise)
                # t.add_column("cxi", list_cxi)
                # t.add_column("cxa", list_cxa)
                # print()
                # print('cya = ' + str(cya))
                # print(t)
        # отрисовка
        for l in list_cxa_POL:
            tck, u = interpolate.splprep([l, list_cya], s=0, k=2)
            xnew, ynew = interpolate.splev(np.linspace(0, 1, 100), tck)
            ax.plot(xnew, ynew, '',  color='k')
            text = 'M = ' + str(list_M[list_cxa_POL.index(l)])
            ax.text(l[-1], list_cya[-1], text, rotation=0, fontsize=7)

        # index = 0
        # for i, j in zip(Cxa_list_up_scrin, cya_list_up_scrin):
        #     ax.annotate(str(a_list_up_scrin[index]), xy=(i, j))
        #     index += 1

        print('ПОЛЕТНЫЕ ПОЛЯРЫ')
        list_H = [0, 3000, 6000, 9000, 12000]
        list_cya_pol = []
        l_H0 = []
        l_H3 = []
        l_H6 = []
        l_H9 = []
        l_H12 = []

        Vmin_rash = math.sqrt((2 * Gpol * g) / (0.85 * pH * S * Cyamax))
        Mmin_rash = Vmin_rash / aH

        row =272
        if type == "ТРД":
            for H_ in list_H:
                self.find_with_H(H_)
                list_cya_pol.clear()
                for M in list_M:
                    cya = 0
                    if M != 0:
                        cya = (2 * Gpol * g) / (pH * S * aH * aH * M * M)
                    else:
                        cya = (2 * Gpol * g) / (pH * S * aH * aH * Mmin_rash * Mmin_rash)
                    # if cya <= 0.7:
                    #     cya = 0.7
                    list_cya_pol.append(float('%.4f' % cya))

                row += 3
                st.cell(row, 2).value = pH
                st.cell(row + 1, 2).value = aH
                row += 3
                i3 = 0
                for col in range(1, 7):
                    cell = st.cell(row, col)
                    cell.value = list_cya_pol[i3]
                    i3 += 1

                l_H = self.Pol_Polar(list_cxa_POL, list_cya_pol)
                list_K.append(self.Find_K(l_H, list_cya_pol))

                tck, u = interpolate.splprep([l_H, list_cya_pol], s=0, k=2)
                xnew, ynew = interpolate.splev(np.linspace(0, 1, 100), tck)
                ax.plot(l_H, list_cya_pol, '.', xnew, ynew, color='tab:blue')
                text = '  H = ' + str(H_)
                ax.text(l_H[-1], list_cya_pol[-1], text, rotation=0, fontsize=7)
                # t = PrettyTable(list_M)
                # t.add_row(list_cya_pol)
                # print()
                # print('H = ' + str(H_))
                # print(t)

        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.set_ylabel('$\it{Cya}$', loc='top', rotation=0)
        ax.set_xlabel('$\it{Cxa}$', loc='right', fontsize=11)
        self.fP_Cre.tight_layout()

        self.cP_Cre.draw()
        self.fP_Cre.savefig('graphics/P_Cre.png')
        wb.save('Расчет самолета.xlsx')

        # except: self.ErrorMessage()

    def Pol_Polar(self, list_cxa_POL, list_cya_pol):
        list_cya = [0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7]
        l_H=[]
        index = -1
        for l in list_cxa_POL:
            tck, u = interpolate.splprep([l, list_cya], s=0, k=2)
            xnew, ynew = interpolate.splev(np.linspace(0, 1, 100), tck)
            min = 1000
            index += 1
            cxa_point = 0
            for x in xnew:
                f = ynew[np.where(xnew == x)] - list_cya_pol[index]
                if 0 < f < min :
                    min = ynew[np.where(xnew == x)] - list_cya_pol[index]
                    cxa_point = x
                if list_cya_pol[index] >= 0.7:
                    cxa_point = l[-1]
            l_H.append(cxa_point)
        return l_H

    def Find_K(self, l_cxa, l_cya):
        list_M = [0, 0.7, 0.8, 0.85, 0.9, 0.95]
        list_K= []
        for M in list_M:
            cxa = l_cxa[list_M.index(M)]
            cya = l_cya[list_M.index(M)]
            try:
                K =cya/cxa
            except: K=0
            list_K.append(float('%.5f' % K))
        return list_K

    def pressedK(self):
        Height = self.main.cb_H.currentText()
        Mah = self.main.cb_M.currentText()
        self.main.la_K.setText(str(self.Show_K(Height, Mah)))

    def Show_K(self, H, M):
        list_H = ['0', '3000', '6000', '9000', '12000']
        list_M = ['0', '0,7', '0,8', '0,85', '0,9', '0,95']
        need_list_K = list_K[list_H.index(H)]
        index = list_M.index(M)
        K = need_list_K[index]
        return K

    def ExportData(self):
        wb = openpyxl.load_workbook('Шаблон ТРД.xlsx') # открытие шаблона
        st_1 = wb['Данные'] # выбор листа для заполнения
        data = [l, S, b, b0, bk, n, c_, xc_, f_, a0, xf_, x_degree, xc_degree, lamda, Sf_,
                Sgd_, Sgsh_, S_, lamdaef, caya, xtau_, cmo, h, bzak_, lzak, Sobzak_, deltavzl,
                deltapos, xshzak, bsrzak, Sobpr_, bgo, cgo_, lgo, Sgo, lamdago, xgo, bv, Sv, ngo,
                bvo, lvo, Svo, cvo_, nvo, bp, cp_, Sp, npylon, lf, Df, Smf, lamdaf, Ssm, lnf,
                lamdanf, lgd, Dgd, lamdagd, Ssm_gd, ln_gd, lamdan_gd, ngd, lgsh, Dgsh, lamdagsh,
                Ssm_gsh, ln_gsh, lamdan_gsh, ngsh, Dv, Fv, Gvzl, V, type, number, P0, H]
        i = 0
        for row in range(1, 82):
            cell = st_1.cell(row, 4) # присвоение значения клетке файла
            if row == 1 or row == 32 or row == 73:
                continue
            else:
                cell.value=data[i]
                i +=1

        # заполнение второго листа с кривыми
        st_2 = wb['Кривые']
        for col in range(2, 10):
            cell = st_2.cell(5, col)
            cell.value = list_Mkr[col-2]

        st_2.cell(7, 2).value = Mrasch
        st_2.cell(8, 2).value = cya_rasch
        st_2.cell(9, 2).value = Gpol

        wb.save('Расчет самолета.xlsx')
        # wb.save('/home/'+getpass.getuser())

    def ExportImage(self, file_info):
        wb = openpyxl.load_workbook('Расчет самолета.xlsx')
        st = wb['Кривые']
        if type == 'ТРД':
            st_2 = wb['Поляры']
        else:
            st_2 = wb['Поляры ТВД']
        img = Image('graphics/Mkr.png') # инициализация изображения
        img.height = 256   # задание размеров
        img.width = 468
        st.add_image(img, 'C10') # присвоение расположения верхнего левого угла изображения
        img_1 = Image('graphics/Help.png')
        img_1.height = 256
        img_1.width = 468
        st.add_image(img_1, 'C34')
        img_2 = Image('graphics/Up.png')
        img_2.height = 322
        img_2.width = 537
        st.add_image(img_2, 'B72')
        img_3 = Image('graphics/Down.png')
        img_3.height = 322
        img_3.width = 537
        st.add_image(img_3, 'B122')
        img_4 = Image('graphics/Cre.png')
        img_4.height = 283
        img_4.width = 537
        st.add_image(img_4, 'B158')
        img_5 = Image('graphics/P_help.png')
        img_5.height = 283
        img_5.width = 537
        st_2.add_image(img_5, 'B31')
        img_5 = Image('graphics/P_Up.png')
        img_5.height = 283
        img_5.width = 537
        st_2.add_image(img_5, 'B72')
        img_5 = Image('graphics/P_Down.png')
        img_5.height = 283
        img_5.width = 537
        st_2.add_image(img_5, 'B118')
        img_8 = Image('graphics/P_Cre.png')
        img_8.height = 283
        img_8.width = 537
        if type == 'ТРД':
            st_2.add_image(img_8, 'B304')
            wb.remove(wb['Поляры ТВД'])
        else:
            st_2.add_image(img_8, 'B258')
            wb.remove(wb['Поляры'])

        wb.save(file_info)

    def ErrorMessage(self):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Critical)
        msg.setText("Убедитесь, что excel файл закрыт.")
        msg.setWindowTitle("Ошибка")
        msg.exec_()

if __name__ == '__main__':  # Если мы запускаем файл напрямую, а не импортируем
    app = QtWidgets.QApplication(sys.argv)  # Новый экземпляр QApplication
    window = ExampleApp()  # Создаём объект класса ExampleApp
    # window.show()  # Показываем окно
    app.exec_()  # и запускаем приложение